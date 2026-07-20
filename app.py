import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from datetime import datetime, timedelta
import json
import requests
from prisma_db import authenticate_user, create_user, save_prediction as prisma_save_prediction, get_user_predictions as prisma_get_user_predictions, get_user_notifications
import prisma_db
import matplotlib.pyplot as plt
import seaborn as sns
from fpdf import FPDF
import base64
from io import BytesIO
import time
from typing import List, Dict, Optional
import hashlib
import hmac
from scipy import stats
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from db import init_db, get_teams_by_sport, get_pitchers, save_future_match
from predict import predict_football, predict_nba, predict_mlb
# ============================================================================
# CONFIGURACIÓN DE LA PÁGINA
# ============================================================================

st.set_page_config(
    page_title="SportsPredict Pro - Predicciones Deportivas",
    page_icon="⚽",
    layout="wide",
    initial_sidebar_state="expanded",
    menu_items={
        "Get Help": "https://www.example.com/help",
        "Report a bug": "https://www.example.com/bug",
        "About": "# SportsPredict Pro v1.0\nPlataforma de predicciones deportivas con IA"
    }
)

# Estilos CSS personalizados
st.markdown("""
    <style>
    .main {
        padding: 2rem;
    }
    .metric-card {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        padding: 1.5rem;
        border-radius: 10px;
        text-align: center;
    }
    .prediction-card {
        background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%);
        color: white;
        padding: 1.5rem;
        border-radius: 10px;
    }
    .premium-badge {
        background: linear-gradient(135deg, #ffd89b 0%, #19547b 100%);
        color: white;
        padding: 0.5rem 1rem;
        border-radius: 20px;
        font-weight: bold;
    }
    </style>
""", unsafe_allow_html=True)

# ============================================================================
# INICIALIZACIÓN Y CONFIGURACIÓN
# ============================================================================

@st.cache_resource
def init_api_keys():
    """Obtiene las claves de API de los secretos (Solo APIs Gratuitas)"""
    return {
        'api_football': st.secrets.get("API_FOOTBALL_KEY", ""),
        'google_studio': st.secrets.get("GOOGLE_STUDIO_KEY", "")
    }

api_keys = init_api_keys()

@st.cache_resource
def init_storage():
    """Asegura que las bases locales existan antes de usarlas."""
    init_db()
    prisma_db.init_prisma_db()

init_storage()

# Estado de sesión
if 'user_id' not in st.session_state:
    st.session_state.user_id = None

if 'user_tier' not in st.session_state:
    st.session_state.user_tier = "free"

if 'user_role' not in st.session_state:
    st.session_state.user_role = "USER"

if 'favorite_teams' not in st.session_state:
    st.session_state.favorite_teams = []

if 'selected_competitions' not in st.session_state:
    st.session_state.selected_competitions = []

if 'theme' not in st.session_state:
    st.session_state.theme = "light"

if 'lang' not in st.session_state:
    st.session_state.lang = "es"

# Función traductora
def t(es: str, en: str, pt: str) -> str:
    lang = st.session_state.get('lang', 'es')
    if lang == 'en':
        return en
    elif lang == 'pt':
        return pt
    return es

# Estilos CSS para el control flotante y modo oscuro dinámico
# Definir e inyectar estilos CSS para el control flotante y modo oscuro dinámico
def inject_theme_css():
    st.markdown("""
        <style>
        .st-key-floating_header {
            position: fixed !important;
            top: 12px !important;
            right: 65px !important;
            z-index: 999999 !important;
            background: """ + ("rgba(30, 41, 59, 0.9)" if st.session_state.theme == "dark" else "rgba(255, 255, 255, 0.9)") + """ !important;
            backdrop-filter: blur(8px) !important;
            padding: 5px 10px !important;
            border-radius: 12px !important;
            box-shadow: """ + ("0 4px 15px rgba(0, 0, 0, 0.4)" if st.session_state.theme == "dark" else "0 4px 15px rgba(0, 0, 0, 0.08)") + """ !important;
            border: 1px solid """ + ("rgba(255, 255, 255, 0.1)" if st.session_state.theme == "dark" else "rgba(0, 0, 0, 0.08)") + """ !important;
            width: auto !important;
        }
        .st-key-floating_header div[data-testid="stHorizontalBlock"] {
            display: flex !important;
            flex-direction: row !important;
            align-items: center !important;
            gap: 6px !important;
        }
        .st-key-floating_header div[data-testid="column"] {
            width: auto !important;
            flex: none !important;
            min-width: 0 !important;
        }
        
        /* Custom Sidebar Styling */
        .sidebar-logo-container {
            display: flex !important;
            align-items: center !important;
            gap: 10px !important;
            padding: 10px 5px !important;
            margin-bottom: 5px !important;
        }
        .sidebar-logo-icon {
            font-size: 1.8rem !important;
        }
        .sidebar-logo-text {
            font-size: 1.35rem !important;
            font-weight: 700 !important;
            letter-spacing: -0.025em !important;
            background: linear-gradient(135deg, #6366f1 0%, #a855f7 100%);
            -webkit-background-clip: text !important;
            -webkit-text-fill-color: transparent !important;
            display: flex !important;
            align-items: center !important;
        }
        .sidebar-logo-text .pro-tag {
            font-size: 0.7rem !important;
            font-weight: 800 !important;
            text-transform: uppercase !important;
            letter-spacing: 0.05em !important;
            background: linear-gradient(135deg, #3b82f6 0%, #1d4ed8 100%) !important;
            -webkit-text-fill-color: white !important;
            color: white !important;
            padding: 2px 6px !important;
            border-radius: 6px !important;
            margin-left: 6px !important;
            display: inline-block !important;
            line-height: 1 !important;
        }
        .sidebar-nav-label {
            font-size: 0.75rem !important;
            font-weight: 700 !important;
            text-transform: uppercase !important;
            letter-spacing: 0.08em !important;
            color: #64748b !important;
            margin-top: 15px !important;
            margin-bottom: 12px !important;
            padding-left: 5px !important;
        }
        
        /* Page links as large cards */
        [data-testid="stSidebar"] [data-testid="stPageLink"] {
            margin: 0 !important;
            padding: 0 !important;
        }
        [data-testid="stSidebar"] [data-testid="stPageLink"] a {
            display: flex !important;
            align-items: center !important;
            padding: 14px 18px !important;
            margin: 8px 0 !important;
            border-radius: 12px !important;
            border: 1px solid var(--sidebar-card-border) !important;
            background-color: var(--sidebar-card-bg) !important;
            color: var(--sidebar-card-fg) !important;
            text-decoration: none !important;
            transition: all 0.25s cubic-bezier(0.4, 0, 0.2, 1) !important;
            box-shadow: 0 1px 2px rgba(0, 0, 0, 0.04) !important;
            height: auto !important;
            width: 100% !important;
        }
        [data-testid="stSidebar"] [data-testid="stPageLink"] a:hover {
            background-color: var(--sidebar-card-hover-bg) !important;
            border-color: var(--sidebar-card-hover-border) !important;
            transform: translateY(-2px) !important;
            box-shadow: 0 6px 16px var(--sidebar-card-hover-shadow) !important;
            color: var(--sidebar-card-hover-fg) !important;
        }
        [data-testid="stSidebar"] [data-testid="stPageLink"] a [data-testid="stIcon"] {
            font-size: 1.3rem !important;
            margin-right: 12px !important;
            display: flex !important;
            align-items: center !important;
            justify-content: center !important;
        }
        [data-testid="stSidebar"] [data-testid="stPageLink"] a p {
            font-size: 0.95rem !important;
            font-weight: 600 !important;
            margin: 0 !important;
            color: inherit !important;
        }
        
        /* Active page card styling using aria-current */
        [data-testid="stSidebar"] [data-testid="stPageLink"] a[aria-current="page"] {
            background: var(--active-gradient) !important;
            border-color: var(--active-border) !important;
            color: #ffffff !important;
            font-weight: 700 !important;
            box-shadow: var(--active-shadow) !important;
        }
        [data-testid="stSidebar"] [data-testid="stPageLink"] a[aria-current="page"] p {
            font-weight: 700 !important;
        }
        
        /* User info card styling */
        .user-info-card {
            background-color: var(--sidebar-card-bg) !important;
            border: 1px solid var(--sidebar-card-border) !important;
            border-radius: 12px !important;
            padding: 14px !important;
            margin-top: 15px !important;
            margin-bottom: 10px !important;
            box-shadow: 0 1px 2px rgba(0, 0, 0, 0.04) !important;
        }
        .user-info-item {
            font-size: 0.825rem !important;
            margin: 4px 0 !important;
            color: var(--sidebar-card-fg) !important;
        }
        </style>
    """, unsafe_allow_html=True)

    if st.session_state.theme == "dark":
        st.markdown("""
            <style>
            .stApp, [data-testid="stAppViewContainer"], [data-testid="stMain"], section[data-testid="stMain"] {
                background-color: #0f172a !important;
                color: #f8fafc !important;
            }
            
            /* Sidebar Dark Theme CSS Override */
            [data-testid="stSidebar"] {
                background-color: #0b0f19 !important;
                border-right: 1px solid #1e293b !important;
                --sidebar-card-bg: #111827 !important;
                --sidebar-card-border: #1f2937 !important;
                --sidebar-card-fg: #9ca3af !important;
                --sidebar-card-hover-bg: #1f2937 !important;
                --sidebar-card-hover-border: #4f46e5 !important;
                --sidebar-card-hover-fg: #ffffff !important;
                --sidebar-card-hover-shadow: rgba(79, 70, 229, 0.25) !important;
                --active-gradient: linear-gradient(135deg, #6366f1 0%, #4f46e5 100%) !important;
                --active-border: #818cf8 !important;
                --active-shadow: 0 4px 12px rgba(99, 102, 241, 0.3) !important;
            }
            [data-testid="stSidebar"] div, [data-testid="stSidebar"] section {
                background-color: transparent !important;
            }
            [data-testid="stSidebar"] * {
                color: #f8fafc !important;
            }
            [data-testid="stSidebar"] button {
                background-color: #1e293b !important;
                color: #f8fafc !important;
                border: 1px solid #334155 !important;
            }
            [data-testid="stSidebar"] button:hover {
                background-color: #334155 !important;
                color: #ffffff !important;
            }
            
            .stTextInput input, .stSelectbox select, div[role="combobox"], select, input {
                background-color: #1e293b !important;
                color: #f8fafc !important;
                border: 1px solid #334155 !important;
                border-radius: 8px !important;
            }
            div[data-testid="stForm"] {
                background-color: #1e293b !important;
                border: 1px solid #334155 !important;
                border-radius: 12px !important;
                padding: 2rem !important;
            }
            .stTextInput input:focus, .stSelectbox select:focus {
                border-color: #6366f1 !important;
                box-shadow: 0 0 0 2px rgba(99, 102, 241, 0.2) !important;
            }
            h1, h2, h3, h4, h5, h6, p, span, li, label, .stMarkdown, .stText, [data-testid="stMetricLabel"], [data-testid="stMetricValue"] {
                color: #f8fafc !important;
            }
            div[data-testid="stExpander"] {
                background-color: #1e293b !important;
                border: 1px solid #334155 !important;
                border-radius: 8px !important;
            }
            .dataframe {
                background-color: #1e293b !important;
                color: #f8fafc !important;
            }
            .metric-card {
                background: linear-gradient(135deg, #4338ca 0%, #6d28d9 100%) !important;
            }
            .prediction-card {
                background: linear-gradient(135deg, #be185d 0%, #be123c 100%) !important;
            }
            .premium-badge {
                background: linear-gradient(135deg, #b45309 0%, #1e3a8a 100%) !important;
            }
            button[data-testid="stBaseButton-secondary"] {
                background-color: #334155 !important;
                color: #f8fafc !important;
                border: 1px solid #475569 !important;
            }
            button[data-testid="stBaseButton-secondary"]:hover {
                background-color: #475569 !important;
                color: #ffffff !important;
            }
            button[data-baseweb="tab"] {
                color: #94a3b8 !important;
            }
            button[data-baseweb="tab"][aria-selected="true"] {
                color: #f8fafc !important;
                border-bottom-color: #6366f1 !important;
            }
            
            /* Dark theme table styling */
            div[data-testid="stTable"] table {
                background-color: #1e293b !important;
                color: #f8fafc !important;
                border-collapse: collapse !important;
                border-radius: 8px !important;
                overflow: hidden !important;
                width: 100% !important;
            }
            div[data-testid="stTable"] th {
                background-color: #334155 !important;
                color: #f8fafc !important;
                border-bottom: 2px solid #475569 !important;
                padding: 10px !important;
            }
            div[data-testid="stTable"] td {
                background-color: #0f172a !important;
                color: #e2e8f0 !important;
                border-bottom: 1px solid #1e293b !important;
                padding: 10px !important;
            }
            /* Dark theme code block styling */
            div[data-testid="stCodeBlock"], pre, code, pre *, code * {
                background-color: #1e293b !important;
                color: #38bdf8 !important; /* cyan for better syntax highlight in dark mode */
            }
            div[data-testid="stCodeBlock"] pre {
                background-color: #1e293b !important;
                border: 1px solid #334155 !important;
                border-radius: 8px !important;
            }
            </style>
        """, unsafe_allow_html=True)
    else:
        st.markdown("""
            <style>
            .stApp, [data-testid="stAppViewContainer"], [data-testid="stMain"], section[data-testid="stMain"] {
                background-color: #ffffff !important;
                color: #1e293b !important;
            }
            
            /* Sidebar Light Theme CSS Override */
            [data-testid="stSidebar"] {
                background-color: #f8fafc !important;
                border-right: 1px solid #e2e8f0 !important;
                --sidebar-card-bg: #ffffff !important;
                --sidebar-card-border: #e2e8f0 !important;
                --sidebar-card-fg: #475569 !important;
                --sidebar-card-hover-bg: #f8fafc !important;
                --sidebar-card-hover-border: #6366f1 !important;
                --sidebar-card-hover-fg: #0f172a !important;
                --sidebar-card-hover-shadow: rgba(99, 102, 241, 0.1) !important;
                --active-gradient: linear-gradient(135deg, #6366f1 0%, #4f46e5 100%) !important;
                --active-border: #4f46e5 !important;
                --active-shadow: 0 4px 12px rgba(99, 102, 241, 0.2) !important;
            }
            [data-testid="stSidebar"] div, [data-testid="stSidebar"] section {
                background-color: transparent !important;
            }
            [data-testid="stSidebar"] * {
                color: #1e293b !important;
            }
            [data-testid="stSidebar"] button {
                background-color: #ffffff !important;
                color: #1e293b !important;
                border: 1px solid #e2e8f0 !important;
            }
            
            div[data-testid="stForm"] {
                background-color: #ffffff !important;
                border: 1px solid #e2e8f0 !important;
                box-shadow: 0 4px 6px -1px rgba(0,0,0,0.05) !important;
            }
            
            /* Light theme table styling */
            div[data-testid="stTable"] table {
                background-color: #ffffff !important;
                color: #1e293b !important;
                border-collapse: collapse !important;
                border-radius: 8px !important;
                overflow: hidden !important;
                width: 100% !important;
            }
            div[data-testid="stTable"] th {
                background-color: #f1f5f9 !important;
                color: #1e293b !important;
                border-bottom: 2px solid #e2e8f0 !important;
                padding: 10px !important;
            }
            div[data-testid="stTable"] td {
                background-color: #ffffff !important;
                color: #334155 !important;
                border-bottom: 1px solid #f1f5f9 !important;
                padding: 10px !important;
            }
            /* Light theme code block styling */
            div[data-testid="stCodeBlock"], pre, code, pre *, code * {
                background-color: #f1f5f9 !important;
                color: #0369a1 !important;
            }
            div[data-testid="stCodeBlock"] pre {
                background-color: #f1f5f9 !important;
                border: 1px solid #e2e8f0 !important;
                border-radius: 8px !important;
            }
            </style>
        """, unsafe_allow_html=True)

# Correr de manera inicial para la pantalla de login
inject_theme_css()

# Renderizado del Control Flotante (Idioma y Tema)
with st.container(key="floating_header"):
    col_lang, col_theme = st.columns([1, 1])
    with col_lang:
        selected_lang = st.selectbox(
            label="Language Select",
            options=["es", "en", "pt"],
            format_func=lambda x: "🇪🇸 ES" if x == "es" else "🇺🇸 EN" if x == "en" else "🇧🇷 PT",
            index=["es", "en", "pt"].index(st.session_state.lang),
            key="lang_select_widget",
            label_visibility="collapsed"
        )
        if selected_lang != st.session_state.lang:
            st.session_state.lang = selected_lang
            st.rerun()
    with col_theme:
        theme_icon = "🌙" if st.session_state.theme == "light" else "☀️"
        if st.button(theme_icon, key="theme_toggle_widget", help="Toggle Light/Dark Theme"):
            st.session_state.theme = "dark" if st.session_state.theme == "light" else "light"
            st.rerun()

# ============================================================================
# CLASES Y UTILIDADES
# ============================================================================

class PDFReport(FPDF):
    """Generador de reportes PDF"""

    def header(self):
        self.set_font('Arial', 'B', 16)
        self.image(None, 10, 8, 33)  # Placeholder para logo
        self.cell(0, 10, 'SportsPredict Pro - Reporte de Predicciones', 0, 1, 'C')
        self.ln(10)

    def footer(self):
        self.set_y(-15)
        self.set_font('Arial', 'I', 8)
        self.cell(0, 10, f'Página {self.page_no()}', 0, 0, 'C')

    def chapter_title(self, title):
        self.set_font('Arial', 'B', 12)
        self.set_fill_color(200, 220, 255)
        self.cell(0, 10, title, 0, 1, 'L', True)
        self.ln(5)

    def chapter_body(self, body):
        self.set_font('Arial', '', 11)
        self.multi_cell(0, 10, body)
        self.ln()

class APIClient:
    """Cliente para consultar APIs de deportes"""

    def __init__(self, api_keys: Dict):
        self.api_keys = api_keys
        self.football_base_url = "https://v3.football.api-sports.io"
        self.session = requests.Session()

    def get_matches_by_date(self, target_date: str) -> List[Dict]:
        """Obtiene partidos de una fecha específica de API-Football"""
        try:
            if not self.api_keys.get('api_football'):
                return self._get_mock_matches()

            headers = {"x-apisports-key": self.api_keys['api_football']}
            response = self.session.get(
                f"{self.football_base_url}/fixtures",
                headers=headers,
                params={"date": target_date, "status": "NS"}  # NS = Not Started
            )
            
            if response.status_code == 200:
                data = response.json()
                errors = data.get('errors')
                if errors:
                    if isinstance(errors, dict) and any(errors.values()):
                        err_msg = ", ".join([f"{k}: {v}" for k, v in errors.items()])
                        raise ValueError(err_msg)
                    elif isinstance(errors, list) and len(errors) > 0:
                        raise ValueError(str(errors))
                
                if 'response' in data:
                    allowed_countries = {'italy', 'spain', 'france', 'germany', 'england'}
                    return [
                        m for m in data['response']
                        if m.get('league', {}).get('country', '').lower() in allowed_countries
                    ]
            else:
                raise ValueError(f"HTTP Status {response.status_code}")
            return self._get_mock_matches()
        except Exception as e:
            st.warning(f"Error obteniendo partidos de API-Football: {e}. Se mostrarán partidos de prueba.")
            return self._get_mock_matches()

    def get_upcoming_matches(self, days: int = 7) -> List[Dict]:
        """Obtiene partidos próximos de API-Football"""
        try:
            if not self.api_keys.get('api_football'):
                return self._get_mock_matches()

            headers = {"x-apisports-key": self.api_keys['api_football']}

            # Obtener partidos de los próximos días
            matches = []
            allowed_countries = {'italy', 'spain', 'france', 'germany', 'england'}
            for i in range(days):
                date = (datetime.now() + timedelta(days=i)).strftime("%Y-%m-%d")
                response = self.session.get(
                    f"{self.football_base_url}/fixtures",
                    headers=headers,
                    params={"date": date, "status": "NS"}  # NS = Not Started
                )

                if response.status_code == 200:
                    data = response.json()
                    errors = data.get('errors')
                    if errors:
                        if isinstance(errors, dict) and any(errors.values()):
                            err_msg = ", ".join([f"{k}: {v}" for k, v in errors.items()])
                            raise ValueError(err_msg)
                        elif isinstance(errors, list) and len(errors) > 0:
                            raise ValueError(str(errors))
                    if 'response' in data:
                        filtered_day = [
                            m for m in data['response']
                            if m.get('league', {}).get('country', '').lower() in allowed_countries
                        ]
                        matches.extend(filtered_day[:5])  # Limitar a 5 por día
                else:
                    raise ValueError(f"HTTP Status {response.status_code}")

            return matches[:10]
        except Exception as e:
            st.warning(f"Error obteniendo partidos de API-Football: {e}. Se mostrarán partidos de prueba.")
            return self._get_mock_matches()

    def get_team_stats(self, team_id: int) -> Dict:
        """Obtiene estadísticas de un equipo"""
        try:
            if not self.api_keys.get('api_football'):
                return self._get_mock_team_stats()

            headers = {"x-apisports-key": self.api_keys['api_football']}
            response = self.session.get(
                f"{self.football_base_url}/teams/statistics",
                headers=headers,
                params={"team": team_id, "season": datetime.now().year}
            )

            if response.status_code == 200:
                data = response.json()
                errors = data.get('errors')
                if errors:
                    if isinstance(errors, dict) and any(errors.values()):
                        err_msg = ", ".join([f"{k}: {v}" for k, v in errors.items()])
                        raise ValueError(err_msg)
                    elif isinstance(errors, list) and len(errors) > 0:
                        raise ValueError(str(errors))
                return data.get('response', {})
            else:
                raise ValueError(f"HTTP Status {response.status_code}")
            return self._get_mock_team_stats()
        except Exception as e:
            st.warning(f"Error obteniendo estadísticas: {e}. Se mostrarán estadísticas de prueba.")
            return self._get_mock_team_stats()

    def get_predictions(self, fixture_id: int) -> Dict:
        """Obtiene predicciones para un partido"""
        try:
            if not self.api_keys.get('api_football'):
                return self._get_mock_prediction()

            headers = {"x-apisports-key": self.api_keys['api_football']}
            response = self.session.get(
                f"{self.football_base_url}/predictions",
                headers=headers,
                params={"fixture": fixture_id}
            )

            if response.status_code == 200:
                data = response.json()
                errors = data.get('errors')
                if errors:
                    if isinstance(errors, dict) and any(errors.values()):
                        err_msg = ", ".join([f"{k}: {v}" for k, v in errors.items()])
                        raise ValueError(err_msg)
                    elif isinstance(errors, list) and len(errors) > 0:
                        raise ValueError(str(errors))
                return data.get('response', [{}])[0]
            else:
                raise ValueError(f"HTTP Status {response.status_code}")
            return self._get_mock_prediction()
        except Exception as e:
            st.warning(f"Error obteniendo predicciones: {e}. Se mostrarán predicciones de prueba.")
            return self._get_mock_prediction()

    def get_head_to_head(self, team1_id: int, team2_id: int) -> List[Dict]:
        """Obtiene historial H2H entre dos equipos"""
        try:
            if not self.api_keys.get('api_football'):
                return self._get_mock_h2h()

            headers = {"x-apisports-key": self.api_keys['api_football']}
            response = self.session.get(
                f"{self.football_base_url}/fixtures/headtohead",
                headers=headers,
                params={"h2h": f"{team1_id}-{team2_id}", "last": 10}
            )

            if response.status_code == 200:
                data = response.json()
                errors = data.get('errors')
                if errors:
                    if isinstance(errors, dict) and any(errors.values()):
                        err_msg = ", ".join([f"{k}: {v}" for k, v in errors.items()])
                        raise ValueError(err_msg)
                    elif isinstance(errors, list) and len(errors) > 0:
                        raise ValueError(str(errors))
                return data.get('response', [])
            else:
                raise ValueError(f"HTTP Status {response.status_code}")
            return self._get_mock_h2h()
        except Exception as e:
            st.warning(f"Error obteniendo H2H: {e}. Se mostrarán partidos de prueba.")
            return self._get_mock_h2h()

    # ==================== DATOS MOCK ====================

    @staticmethod
    def _get_mock_matches() -> List[Dict]:
        """Retorna datos de ejemplo para partidos"""
        return [
            {
                'fixture': {'id': 1001, 'date': '2026-07-13T20:00:00+00:00', 'status': 'NS'},
                'league': {'name': 'La Liga', 'country': 'Spain', 'logo': ''},
                'teams': {
                    'home': {'id': 541, 'name': 'Real Madrid', 'logo': ''},
                    'away': {'id': 529, 'name': 'Barcelona', 'logo': ''}
                },
                'goals': {'home': None, 'away': None}
            },
            {
                'fixture': {'id': 1002, 'date': '2026-07-13T15:00:00+00:00', 'status': 'NS'},
                'league': {'name': 'Premier League', 'country': 'England', 'logo': ''},
                'teams': {
                    'home': {'id': 42, 'name': 'Manchester City', 'logo': ''},
                    'away': {'id': 40, 'name': 'Liverpool', 'logo': ''}
                },
                'goals': {'home': None, 'away': None}
            },
            {
                'fixture': {'id': 1003, 'date': '2026-07-13T19:30:00+00:00', 'status': 'NS'},
                'league': {'name': 'Bundesliga', 'country': 'Germany', 'logo': ''},
                'teams': {
                    'home': {'id': 25, 'name': 'Bayern Munich', 'logo': ''},
                    'away': {'id': 165, 'name': 'Borussia Dortmund', 'logo': ''}
                },
                'goals': {'home': None, 'away': None}
            },
            {
                'fixture': {'id': 1004, 'date': '2026-07-13T18:00:00+00:00', 'status': 'NS'},
                'league': {'name': 'La Liga', 'country': 'Spain', 'logo': ''},
                'teams': {
                    'home': {'id': 542, 'name': 'Atletico Madrid', 'logo': ''},
                    'away': {'id': 529, 'name': 'Barcelona', 'logo': ''}
                },
                'goals': {'home': None, 'away': None}
            },
            {
                'fixture': {'id': 1005, 'date': '2026-07-13T21:00:00+00:00', 'status': 'NS'},
                'league': {'name': 'La Liga', 'country': 'Spain', 'logo': ''},
                'teams': {
                    'home': {'id': 548, 'name': 'Real Sociedad', 'logo': ''},
                    'away': {'id': 541, 'name': 'Real Madrid', 'logo': ''}
                },
                'goals': {'home': None, 'away': None}
            },
            {
                'fixture': {'id': 1006, 'date': '2026-07-13T17:30:00+00:00', 'status': 'NS'},
                'league': {'name': 'Premier League', 'country': 'England', 'logo': ''},
                'teams': {
                    'home': {'id': 49, 'name': 'Arsenal', 'logo': ''},
                    'away': {'id': 45, 'name': 'Chelsea', 'logo': ''}
                },
                'goals': {'home': None, 'away': None}
            },
            {
                'fixture': {'id': 1007, 'date': '2026-07-13T20:00:00+00:00', 'status': 'NS'},
                'league': {'name': 'Premier League', 'country': 'England', 'logo': ''},
                'teams': {
                    'home': {'id': 33, 'name': 'Manchester United', 'logo': ''},
                    'away': {'id': 47, 'name': 'Tottenham', 'logo': ''}
                },
                'goals': {'home': None, 'away': None}
            },
            {
                'fixture': {'id': 1008, 'date': '2026-07-13T15:30:00+00:00', 'status': 'NS'},
                'league': {'name': 'Bundesliga', 'country': 'Germany', 'logo': ''},
                'teams': {
                    'home': {'id': 168, 'name': 'Bayer Leverkusen', 'logo': ''},
                    'away': {'id': 25, 'name': 'Bayern Munich', 'logo': ''}
                },
                'goals': {'home': None, 'away': None}
            },
            {
                'fixture': {'id': 1009, 'date': '2026-07-13T18:30:00+00:00', 'status': 'NS'},
                'league': {'name': 'Bundesliga', 'country': 'Germany', 'logo': ''},
                'teams': {
                    'home': {'id': 173, 'name': 'RB Leipzig', 'logo': ''},
                    'away': {'id': 165, 'name': 'Borussia Dortmund', 'logo': ''}
                },
                'goals': {'home': None, 'away': None}
            }
        ]

    @staticmethod
    def _get_mock_prediction() -> Dict:
        """Retorna predicción de ejemplo"""
        return {
            'predictions': {
                'winner': {'name': 'home', 'comment': 'The home team is likely to win'},
                'win_or_draw': 95,
                'win_home_or_draw': 88,
                'goals': {'home': 2.1, 'away': 1.2},
                'goals_more_less': 'over',
                'advice': 'Home team is in better form'
            },
            'comparison': {
                'form': {'home': '5W-2D-1L', 'away': '3W-4D-3L'},
                'att': {'home': 85, 'away': 72},
                'def': {'home': 92, 'away': 78},
                'poisson_distribution': {
                    'home': {'0': 8, '1': 21, '2': 35, '3': 30},
                    'away': {'0': 15, '1': 34, '2': 28, '3': 16}
                }
            }
        }

    @staticmethod
    def _get_mock_team_stats() -> Dict:
        """Retorna estadísticas de equipo de ejemplo"""
        return {
            'team': {'id': 541, 'name': 'Real Madrid'},
            'statistics': [
                {'type': 'Wins home', 'value': 12},
                {'type': 'Wins away', 'value': 8},
                {'type': 'Draws', 'value': 4},
                {'type': 'Goals for', 'value': 65},
                {'type': 'Goals against', 'value': 18},
                {'type': 'Goals for home', 'value': 35},
                {'type': 'Goals for away', 'value': 30},
                {'type': 'Goals against home', 'value': 8},
                {'type': 'Goals against away', 'value': 10}
            ]
        }

    @staticmethod
    def _get_mock_h2h() -> List[Dict]:
        """Retorna historial H2H de ejemplo"""
        return [
            {
                'fixture': {'id': 900, 'date': '2023-12-10T19:00:00+00:00'},
                'teams': {'home': {'name': 'Real Madrid'}, 'away': {'name': 'Barcelona'}},
                'goals': {'home': 2, 'away': 1},
                'score': {'fulltime': {'home': 2, 'away': 1}}
            },
            {
                'fixture': {'id': 901, 'date': '2023-10-28T20:00:00+00:00'},
                'teams': {'home': {'name': 'Barcelona'}, 'away': {'name': 'Real Madrid'}},
                'goals': {'home': 1, 'away': 1},
                'score': {'fulltime': {'home': 1, 'away': 1}}
            },
            {
                'fixture': {'id': 902, 'date': '2023-08-15T19:30:00+00:00'},
                'teams': {'home': {'name': 'Real Madrid'}, 'away': {'name': 'Barcelona'}},
                'goals': {'home': 3, 'away': 2},
                'score': {'fulltime': {'home': 3, 'away': 2}}
            }
        ]

api_client = APIClient(api_keys)

# ============================================================================
# FUNCIONES DE BASE DE DATOS
# ============================================================================

def save_prediction(user_id: str, match_data: Dict, prediction: Dict, confidence: float):
    """Guarda una predicción en la base de datos"""
    try:
        match_id = match_data.get('fixture', {}).get('id', 0)
        # Using prisma to save
        prisma_save_prediction(
            user_id=user_id,
            match_id=match_id,
            predicted_winner=prediction.get('winner', {}).get('name', 'Unknown'),
            confidence=confidence,
            is_manual=False,
            ai_data=json.dumps(prediction)
        )
        return True
    except Exception as e:
        st.error(f"Error guardando predicción: {e}")
        return False

def get_user_predictions(user_id: str, limit: int = 50) -> List[Dict]:
    """Obtiene las predicciones de un usuario"""
    try:
        if not user_id: return []
        preds = prisma_get_user_predictions(user_id, limit)
        # Map back to dict format expected by the app
        mapped = []
        for p in preds:
            mapped.append({
                'match_id': p.get('matchId'),
                'predicted_home_score': p.get('predictedHomeScore', 0),
                'predicted_away_score': p.get('predictedAwayScore', 0),
                'confidence_level': p.get('confidenceLevel', 0),
                'created_at': p.get('createdAt').isoformat() if p.get('createdAt') else "",
                'is_manual': p.get('isManual', False),
                'predicted_winner': p.get('predictedWinner', '')
            })
        return mapped
    except Exception as e:
        st.warning(f"Error obteniendo predicciones: {e}")
        return []

def get_user_stats(user_id: str) -> Dict:
    """Calcula estadísticas del usuario"""
    predictions = get_user_predictions(user_id)

    if not predictions:
        return {
            'total_predictions': 0,
            'correct_predictions': 0,
            'accuracy_rate': 0,
            'avg_confidence': 0,
            'total_earnings': 0,
            'rank': 'N/A'
        }

    df = pd.DataFrame(predictions)

    total_preds = len(df)
    correct_preds = len(df[df['confidence_level'] > 0.7]) if 'confidence_level' in df.columns else 0
    avg_conf = df['confidence_level'].mean() if 'confidence_level' in df.columns else 0

    return {
        'total_predictions': total_preds,
        'correct_predictions': correct_preds,
        'accuracy_rate': (correct_preds / total_preds * 100) if total_preds > 0 else 0,
        'avg_confidence': float(avg_conf),
        'total_earnings': 0,
        'rank': f'Top {np.random.randint(5, 25)}%'
    }

def get_competitions() -> List[Dict]:
    """Obtiene las competencias activas"""
    return []


def is_admin() -> bool:
    """Indica si el usuario autenticado tiene rol de administrador."""
    return st.session_state.get('user_role', 'USER') == 'ADMIN'

# ============================================================================
# COMPONENTES REUTILIZABLES
# ============================================================================

def render_metric_card(label: str, value: str, delta: str = "", icon: str = ""):
    """Renderiza una tarjeta de métrica"""
    col1, col2, col3 = st.columns([1, 2, 1])
    with col1:
        st.markdown(f"<h3>{icon}</h3>", unsafe_allow_html=True)
    with col2:
        st.metric(label, value, delta=delta)
    with col3:
        pass

def render_match_card(match: Dict):
    """Renderiza una tarjeta de partido"""
    try:
        fixture = match.get('fixture', {})
        teams = match.get('teams', {})
        league = match.get('league', {})

        home_team = teams.get('home', {})
        away_team = teams.get('away', {})

        match_date = fixture.get('date', '')
        if match_date:
            match_date = datetime.fromisoformat(match_date.replace('Z', '+00:00')).strftime("%d/%m/%Y %H:%M")

        col1, col2, col3, col4 = st.columns([2, 2, 2, 2])

        with col1:
            st.write(f"🏆 **{league.get('name', 'Liga')}**")
            st.write(f"📅 {match_date}")

        with col2:
            st.write(f"🏠 **{home_team.get('name', 'Equipo 1')}**")

        with col3:
            st.write("vs")

        with col4:
            st.write(f"✈️ **{away_team.get('name', 'Equipo 2')}**")

        return True
    except Exception as e:
        st.error(f"Error renderizando partido: {e}")
        return False

# ============================================================================
# GENERADOR DE REPORTES - PDF Y EXCEL
# ============================================================================

def generate_predictions_report_pdf(predictions: List[Dict]) -> bytes:
    """Genera reporte PDF de predicciones"""
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", "B", 16)
    pdf.cell(0, 15, "REPORTE DE PREDICCIONES", ln=True, align="C")

    pdf.set_font("Arial", "", 11)
    pdf.ln(5)

    df = pd.DataFrame(predictions) if predictions else pd.DataFrame()

    if not df.empty:
        total = len(df)
        correct = len(df[df.get('prediction_status') == 'won']) if 'prediction_status' in df.columns else 0
        accuracy = (correct / total * 100) if total > 0 else 0

        pdf.cell(0, 8, f"Total de Predicciones: {total}", ln=True)
        pdf.cell(0, 8, f"Predicciones Correctas: {correct}", ln=True)
        pdf.cell(0, 8, f"Tasa de Precisión: {accuracy:.2f}%", ln=True)

        if 'confidence_level' in df.columns:
            avg_conf = df['confidence_level'].mean()
            pdf.cell(0, 8, f"Confianza Promedio: {avg_conf:.2f}", ln=True)

    pdf.ln(5)
    pdf.set_font("Arial", "I", 8)
    timestamp = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    pdf.cell(0, 10, f"Generado: {timestamp}", align="R")

    return pdf.output()


def generate_predictions_report_excel(predictions: List[Dict]) -> bytes:
    """Genera reporte Excel de predicciones"""
    df = pd.DataFrame(predictions) if predictions else pd.DataFrame()

    wb = Workbook()
    ws = wb.active
    ws.title = "Predicciones"

    # Headers
    if not df.empty:
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = col_name
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")

        # Datos
        for row_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ============================================================================
# ESTADÍSTICAS DESCRIPTIVAS - GRÁFICOS Y ANÁLISIS
# ============================================================================

def create_distribution_plot(data: List[float], title: str = "Distribución") -> go.Figure:
    """Crea gráfico de distribución con curva normal"""
    if not data:
        return go.Figure()

    df = pd.DataFrame({'value': data})

    fig = go.Figure()

    fig.add_trace(go.Histogram(
        x=df['value'],
        name='Frecuencia',
        nbinsx=30,
        marker_color='rgba(102, 126, 234, 0.7)'
    ))

    mu, sigma = np.mean(data), np.std(data)
    x = np.linspace(min(data), max(data), 100)
    y = stats.norm.pdf(x, mu, sigma)
    y_scaled = y * len(data) * (max(data) - min(data)) / 30

    fig.add_trace(go.Scatter(
        x=x, y=y_scaled,
        name='Curva Normal',
        mode='lines',
        line=dict(color='red', width=3)
    ))

    fig.update_layout(
        title=title,
        xaxis_title="Valor",
        yaxis_title="Frecuencia",
        hovermode='x unified',
        height=400
    )

    return fig


def create_boxplot(data_dict: Dict[str, List[float]], title: str = "Comparación") -> go.Figure:
    """Crea gráfico de caja (boxplot)"""
    fig = go.Figure()

    for label, values in data_dict.items():
        fig.add_trace(go.Box(y=values, name=label, boxmean='sd'))

    fig.update_layout(
        title=title,
        yaxis_title="Valor",
        height=400
    )

    return fig


def create_time_series_plot(dates: List[datetime], values: List[float], title: str = "Serie Temporal") -> go.Figure:
    """Crea gráfico de serie temporal"""
    df = pd.DataFrame({'date': dates, 'value': values})
    df = df.sort_values('date')

    fig = go.Figure()

    fig.add_trace(go.Scatter(
        x=df['date'],
        y=df['value'],
        mode='lines+markers',
        name='Valor',
        line=dict(color='#667EEA', width=2)
    ))

    if len(df) > 7:
        df['ma_7'] = df['value'].rolling(window=7).mean()
        fig.add_trace(go.Scatter(
            x=df['date'],
            y=df['ma_7'],
            mode='lines',
            name='Media Móvil (7 días)',
            line=dict(color='red', dash='dash')
        ))

    fig.update_layout(
        title=title,
        xaxis_title="Fecha",
        yaxis_title="Valor",
        height=400
    )

    return fig


def create_correlation_heatmap(df: pd.DataFrame, numeric_cols: List[str]) -> go.Figure:
    """Crea mapa de calor de correlaciones"""
    numeric_df = df[numeric_cols].select_dtypes(include=[np.number])
    corr_matrix = numeric_df.corr()

    fig = go.Figure(data=go.Heatmap(
        z=corr_matrix.values,
        x=corr_matrix.columns,
        y=corr_matrix.columns,
        colorscale='RdBu',
        zmid=0,
        text=np.round(corr_matrix.values, 2),
        texttemplate='%{text:.2f}'
    ))

    fig.update_layout(title="Matriz de Correlación", height=500, width=600)
    return fig


def calculate_descriptive_stats(data: List[float]) -> Dict:
    """Calcula estadísticas descriptivas"""
    if not data or len(data) == 0:
        return {}

    data_array = np.array(data)

    return {
        'count': len(data),
        'mean': float(np.mean(data_array)),
        'median': float(np.median(data_array)),
        'std': float(np.std(data_array)),
        'min': float(np.min(data_array)),
        'max': float(np.max(data_array)),
        'q1': float(np.percentile(data_array, 25)),
        'q3': float(np.percentile(data_array, 75))
    }


def create_descriptive_stats_table(predictions: List[Dict]) -> pd.DataFrame:
    """Crea tabla de estadísticas descriptivas"""
    if not predictions:
        return pd.DataFrame()

    df = pd.DataFrame(predictions)
    numeric_cols = df.select_dtypes(include=[np.number]).columns

    stats_data = []

    for col in numeric_cols:
        col_stats = calculate_descriptive_stats(df[col].dropna().tolist())
        stats_data.append({
            'Variable': col,
            'N': col_stats.get('count', 0),
            'Media': round(col_stats.get('mean', 0), 2),
            'Mediana': round(col_stats.get('median', 0), 2),
            'Desv. Est.': round(col_stats.get('std', 0), 2),
            'Mín': round(col_stats.get('min', 0), 2),
            'Máx': round(col_stats.get('max', 0), 2)
        })

    return pd.DataFrame(stats_data)

# ============================================================================
# PÁGINAS PRINCIPALES
# ============================================================================

def page_dashboard():
    """Página principal - Dashboard"""
    st.title(t("⚽ Dashboard de Predicciones Deportivas", "⚽ Sports Predictions Dashboard", "⚽ Dashboard de Previsões Esportivas"))

    # Métricas principales
    st.subheader(t("Tus Estadísticas", "Your Statistics", "Suas Estatísticas"))

    user_stats = get_user_stats(st.session_state.user_id)

    col1, col2, col3, col4 = st.columns(4)

    with col1:
        st.metric(
            t("Total Predicciones", "Total Predictions", "Total de Previsões"),
            user_stats['total_predictions'],
            delta=f"{np.random.randint(1, 5)} " + t("hoy", "today", "hoje")
        )

    with col2:
        st.metric(
            t("Tasa de Precisión", "Accuracy Rate", "Taxa de Precisão"),
            f"{user_stats['accuracy_rate']:.1f}%",
            delta=f"+{np.random.randint(1, 5)}%" if user_stats['accuracy_rate'] > 0 else "0%"
        )

    with col3:
        st.metric(
            t("Confianza Promedio", "Average Confidence", "Confiança Média"),
            f"{user_stats['avg_confidence']:.2f}",
            delta="+0.05" if user_stats['avg_confidence'] > 0 else "0.00"
        )

    with col4:
        st.metric(
            t("Ranking Global", "Global Rank", "Classificação Global"),
            user_stats['rank'],
            delta=f"+5 " + t("posiciones", "positions", "posições")
        )

    st.divider()

    # Predicciones recientes desde la base de datos Neon (PostgreSQL)
    st.subheader(t("📋 Mis Predicciones Recientes (Cloud)", "📋 My Recent Predictions (Cloud)", "📋 Minhas Previsões Recentes (Cloud)"))
    st.markdown(t("Tus últimas predicciones guardadas y sincronizadas en la nube.", "Your latest predictions saved and synchronized in the cloud.", "Suas últimas previsões salvas e sincronizadas na nuvem."))

    with st.spinner(t("Cargando predicciones desde Neon...", "Loading predictions from Neon...", "Carregando previsões do Neon...")):
        recent_preds = get_user_predictions(st.session_state.user_id, limit=5)

    if recent_preds:
        for i, pred in enumerate(recent_preds):
            try:
                # Intentar parsear la fecha de creación
                date_obj = datetime.fromisoformat(pred['created_at'].replace('Z', '+00:00'))
                date_str = date_obj.strftime("%d/%m/%Y %H:%M")
            except Exception:
                date_str = pred['created_at']

            is_manual_str = t("Manual", "Manual", "Manual") if pred['is_manual'] else t("Automática (IA)", "Automatic (AI)", "Automática (IA)")
            badge_color = "#3b82f6" if pred['is_manual'] else "#10b981"
            
            with st.expander(f"🔮 Partido ID: {pred['match_id']} — {date_str}"):
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.markdown(t("**Resultado Predicho:**", "**Predicted Result:**", "**Resultado Previsto:**"))
                    st.info(f"🏆 {pred['predicted_winner']}")
                with col2:
                    st.markdown(t("**Nivel de Confianza:**", "**Confidence Level:**", "**Nível de Confiança:**"))
                    st.metric(label=t("Confianza", "Confidence", "Confiança"), value=f"{pred['confidence_level']*100:.1f}%")
                with col3:
                    st.markdown(t("**Detalles del Registro:**", "**Record Details:**", "**Detalhes do Registro:**"))
                    st.markdown(f"**Tipo:** <span style='color:{badge_color};font-weight:bold;'>{is_manual_str}</span>", unsafe_allow_html=True)
                    st.markdown(t(f"**Fecha:** {date_str}", f"**Date:** {date_str}", f"**Data:** {date_str}"))
    else:
        st.info(t("Aún no tienes predicciones registradas en la nube. Ve a 'Crear Predicción' o 'Predicciones IA' para generar una.",
                  "You don't have any predictions registered in the cloud yet. Go to 'Create Prediction' or 'AI Predictions' to generate one.",
                  "Você ainda não tem previsões registradas na nuvem. Vá para 'Criar Previsão' ou 'Previsões IA' para gerar uma."))

    st.divider()

    # Gráfico de rendimiento
    st.subheader(t("📈 Tu Rendimiento Reciente", "📈 Your Recent Performance", "📈 Seu Desempenho Recente"))

    # Datos simulados
    days = pd.date_range(end=datetime.now(), periods=30)
    df_performance = pd.DataFrame({
        'fecha': days,
        'precisión': np.random.uniform(0.4, 0.85, 30),
        'predicciones': np.random.randint(1, 8, 30)
    })

    col1, col2 = st.columns(2)

    with col1:
        fig_precision = px.line(
            df_performance,
            x='fecha',
            y='precisión',
            title=t('Evolución de Precisión (30 días)', 'Accuracy Evolution (30 days)', 'Evolução da Precisão (30 dias)'),
            markers=True,
            color_discrete_sequence=['#FF4B4B']
        )
        fig_precision.update_layout(hovermode='x unified')
        st.plotly_chart(fig_precision, use_container_width=True)

    with col2:
        fig_volume = px.bar(
            df_performance,
            x='fecha',
            y='predicciones',
            title=t('Volumen de Predicciones', 'Prediction Volume', 'Volume de Previsões'),
            color_discrete_sequence=['#00CC96']
        )
        st.plotly_chart(fig_volume, use_container_width=True)

def page_my_predictions():
    """Página de mis predicciones"""
    st.title(t("📊 Mis Predicciones", "📊 My Predictions", "📊 Minhas Previsões"))

    predictions = get_user_predictions(st.session_state.user_id)

    if predictions:
        df = pd.DataFrame(predictions)

        # Resumen
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric(t("Total de predicciones", "Total predictions", "Total de previsões"), len(df))
        with col2:
            high_conf = len(df[df['confidence_level'] > 0.7])
            st.metric(t("Alta confianza (>70%)", "High confidence (>70%)", "Alta confiança (>70%)"), high_conf)
        with col3:
            avg_conf = df['confidence_level'].mean()
            st.metric(t("Confianza promedio", "Average confidence", "Confiança média"), f"{avg_conf:.2f}")

        st.divider()

        # Tabla de predicciones
        st.subheader(t("Historial de Predicciones", "Prediction History", "Histórico de Previsões"))

        display_df = df[['match_id', 'predicted_home_score', 'predicted_away_score', 'confidence_level', 'created_at']].copy()
        display_df.columns = [
            t('ID Partido', 'Match ID', 'ID do Jogo'),
            t('Goles Local', 'Home Goals', 'Gols do Local'),
            t('Goles Visitante', 'Away Goals', 'Gols do Visitante'),
            t('Confianza', 'Confidence', 'Confiança'),
            t('Fecha', 'Date', 'Data')
        ]

        st.table(display_df)

        # Análisis
        st.subheader(t("📈 Análisis de Confianza", "📈 Confidence Analysis", "📈 Análise de Confiança"))

        fig = px.histogram(
            df,
            x='confidence_level',
            nbins=20,
            title=t('Distribución de Niveles de Confianza', 'Distribution of Confidence Levels', 'Distribuição dos Níveis de Confiança'),
            color_discrete_sequence=['#FF4B4B']
        )
        st.plotly_chart(fig, use_container_width=True)
    else:
        st.info(t("Aún no has realizado predicciones. ¡Comienza ahora!", "You haven't made any predictions yet. Start now!", "Você ainda não fez nenhuma previsão. Comece agora!"))

import json
from reports import generate_consolidated_report

def page_ai_predictions():
    """Página de predicciones con IA"""
    st.title(t("🤖 Predicciones con Inteligencia Artificial", "🤖 Artificial Intelligence Predictions", "🤖 Previsões com Inteligência Artificial"))

    st.markdown(t(
        "Nuestra inteligencia artificial se somete a una rigurosa evaluación de validación cruzada. Se utiliza el **historial histórico completo** para entrenamiento, reservando estrictamente los **últimos 365 días** de resultados reales para evaluar la precisión del modelo en escenarios no vistos.",
        "Our artificial intelligence undergoes rigorous cross-validation evaluation. The **full historical records** are used for training, strictly reserving the **last 365 days** of real results to evaluate model accuracy in unseen scenarios.",
        "Nossa inteligência artificial passa por uma rigorosa avaliação de validação cruzada. O **historico completo** é utilizado para treinamento, reservando estritamente os **últimos 365 dias** de resultados reais para avaliar a precisão del modelo em cenarios não vistos."
    ))

    st.divider()

    # Inicializar el estado de la predicción activa
    if 'active_prediction' not in st.session_state:
        st.session_state.active_prediction = None

    # Mostrar la última predicción realizada si está activa
    if st.session_state.active_prediction:
        act = st.session_state.active_prediction
        box_bg = "#1e293b" if st.session_state.theme == "dark" else "#f8fafc"
        box_text = "#f8fafc" if st.session_state.theme == "dark" else "#1e293b"
        st.markdown(f"<div style='border: 2px solid #667eea; padding: 20px; border-radius: 12px; margin-bottom: 20px; background-color: {box_bg}; color: {box_text};'>", unsafe_allow_html=True)
        st.subheader(t("🔮 Último Análisis de IA Generado", "🔮 Latest AI Analysis Generated", "🔮 Última Análise de IA Gerada"))
        col_m_left, col_m_right = st.columns([3, 2])
        with col_m_left:
            st.write(f"### {act['home_team']} vs {act['away_team']}")
            st.success(t(f"Favorito: **{act['pred_label'].upper()}**", f"Favorite: **{act['pred_label'].upper()}**", f"Favorito: **{act['pred_label'].upper()}**"))
            st.info(t(f"📈 **Confianza:** {act['confidence']*100:.1f}%", f"📈 **Confidence:** {act['confidence']*100:.1f}%", f"📈 **Confiança:** {act['confidence']*100:.1f}%"))
        with col_m_right:
            labels = [t('Visitante', 'Away', 'Visitante'), t('Empate', 'Draw', 'Empate'), t('Local', 'Home', 'Local')]
            values = [act['probs']['visitante']*100, act['probs']['empate']*100, act['probs']['local']*100]
            fig = go.Figure(data=[go.Pie(labels=labels, values=values, hole=.5, marker_colors=['#f5576c', '#aaaaaa', '#667eea'])])
            fig.update_layout(height=220, margin=dict(t=20, b=0, l=0, r=0))
            st.plotly_chart(fig, use_container_width=True)
        
        if st.button(t("❌ Cerrar Análisis", "❌ Close Analysis", "❌ Fechar Análise"), use_container_width=True):
            st.session_state.active_prediction = None
            st.rerun()
        st.markdown("</div>", unsafe_allow_html=True)

    st.subheader(t("📅 Partidos Futuros (Mundial)", "📅 Future Matches (Worldwide)", "📅 Jogos Futuros (Mundial)"))
    st.write(t("Selecciona una fecha (hasta 4 días en el futuro) para ver los próximos partidos de fútbol y obtener una predicción de nuestra IA.", "Select a date (up to 4 days in the future) to see upcoming soccer matches and get an AI prediction.", "Selecione uma data (até 4 dias no futuro) para ver os próximos jogos de futebol e obter uma previsão da nossa IA."))
    
    if 'api_matches' not in st.session_state:
        st.session_state.api_matches = []
        st.session_state.last_api_date = None
        
    col_date, col_empty = st.columns([1, 3])
    with col_date:
        selected_date = st.date_input(
            t("Fecha de los partidos", "Matches date", "Data dos jogos"),
            value=datetime.now().date(),
            min_value=datetime.now().date(),
            max_value=(datetime.now() + timedelta(days=4)).date()
        )
    
    date_str = selected_date.strftime("%Y-%m-%d")
    
    if st.session_state.last_api_date != date_str:
        with st.spinner(t("Conectando con API-Football...", "Connecting to API-Football...", "Conectando com a API-Football...")):
            st.session_state.api_matches = api_client.get_matches_by_date(date_str)
            st.session_state.last_api_date = date_str

    # Obtener IDs de partidos ya predichos por el usuario
    user_preds = get_user_predictions(st.session_state.user_id)
    predicted_match_ids = {int(p['match_id']) for p in user_preds if p.get('match_id') is not None}
            
    if st.session_state.api_matches:
        # Filtrar partidos que ya fueron predichos
        matches = [m for m in st.session_state.api_matches if m.get('fixture', {}).get('id') not in predicted_match_ids]
        
        if matches:
            st.success(t(f"Se encontraron {len(matches)} partidos programados para el {st.session_state.last_api_date} (excluyendo tus predicciones).",
                         f"Found {len(matches)} matches scheduled for {st.session_state.last_api_date} (excluding your predictions).",
                         f"Foram encontrados {len(matches)} jogos agendados para {st.session_state.last_api_date} (excluindo suas previsões)."))
            
            for i, match in enumerate(matches[:15]): 
                fixture = match.get('fixture', {})
                teams = match.get('teams', {})
                league = match.get('league', {})
                
                home_team = teams.get('home', {}).get('name', t('Local', 'Home', 'Local'))
                away_team = teams.get('away', {}).get('name', t('Visitante', 'Away', 'Visitante'))
                match_time = fixture.get('date', '').split('T')[1][:5] if 'T' in fixture.get('date', '') else ''
                
                with st.container():
                    box_bg = "#1e293b" if st.session_state.theme == "dark" else "#f8fafc"
                    box_border = "#334155" if st.session_state.theme == "dark" else "#e2e8f0"
                    box_text = "#f8fafc" if st.session_state.theme == "dark" else "#1e293b"
                    st.markdown(f"<div style='border: 1px solid {box_border}; padding: 15px; border-radius: 8px; margin-bottom: 10px; background-color: {box_bg}; color: {box_text};'>", unsafe_allow_html=True)
                    col1, col2, col3 = st.columns([2, 3, 2])
                    with col1:
                        st.write(t(f"🏆 **{league.get('name', 'Liga')}**", f"🏆 **{league.get('name', 'League')}**", f"🏆 **{league.get('name', 'Liga')}**"))
                        st.write(f"⏱️ {match_time}")
                    with col2:
                        st.write(f"🏠 **{home_team}**")
                        st.write(f"✈️ **{away_team}**")
                    
                    with col3:
                        # Usar el ID del fixture en la clave del botón para evitar colisiones
                        btn_key = f"pred_api_{fixture.get('id', i)}"
                        if st.button(t("🔮 Predecir", "🔮 Predict", "🔮 Prever"), key=btn_key, use_container_width=True):
                            with st.spinner(t("Analizando historial...", "Analyzing history...", "Analisando histórico...")):
                                try:
                                    probs = predict_football(home_team, away_team)
                                    pred_label = max(probs, key=probs.get)
                                    
                                    # Guardar en la base de datos de Prisma del usuario
                                    fid = fixture.get('id') if fixture.get('id') else int(time.time())
                                    prisma_save_prediction(
                                        user_id=st.session_state.user_id,
                                        match_id=fid,
                                        predicted_winner=pred_label.upper(),
                                        confidence=max(probs.values()),
                                        is_manual=False,
                                        ai_data=json.dumps({"prediccion_ia": pred_label, "probs": probs})
                                    )
                                    
                                    # Guardar en session state para el banner
                                    st.session_state.active_prediction = {
                                        'home_team': home_team,
                                        'away_team': away_team,
                                        'pred_label': pred_label,
                                        'probs': probs,
                                        'confidence': max(probs.values())
                                    }
                                    st.rerun()
                                except Exception as e:
                                    st.error(f"Error: {str(e)}")
                    st.markdown("</div>", unsafe_allow_html=True)
            
            if len(matches) > 15:
                st.info(t(f"... y {len(matches) - 15} partidos más. (Se muestran los primeros 15).", f"... and {len(matches) - 15} more matches. (Showing first 15).", f"... e mais {len(matches) - 15} jogos. (Mostrando os primeiros 15)."))
        else:
            st.info(t("Ya has realizado predicciones para todos los partidos disponibles en esta fecha.", "You have already made predictions for all available matches on this date.", "Você já fez previsões para todos os jogos disponíveis nesta data."))
            
    st.divider()

    if not is_admin():
        st.info(t("Los resultados detallados del modelo y la exportación consolidada están restringidos al rol ADMIN.", "Detailed model results and consolidated export are restricted to the ADMIN role.", "Resultados detalhados do modelo e exportação consolidada são restritos à função ADMIN."))
        return

    try:
        with open('model_results.json', 'r') as f:
            model_results = json.load(f)
    except FileNotFoundError:
        st.warning(t("⏳ Los modelos se están entrenando en este momento (Time-Based Split). Por favor, regresa más tarde para ver los resultados.", "⏳ Models are currently training (Time-Based Split). Please return later to see the results.", "⏳ Os modelos estão treinando no momento (Time-Based Split). Por favor, volte mais tarde para ver os resultados."))
        return

    st.subheader("📊 Resultados de Validación Cronológica (Último Año)")
    
    # Selector de deporte
    deportes = list(model_results.keys())
    sport = st.selectbox(t("Deporte a visualizar", "Sport to visualize", "Esporte a visualizar"), deportes, format_func=lambda x: x.upper())
    
    result = model_results[sport]
    best_model = result['best_model']
    model_details = result['model_details']

    st.info(t(f"🏆 El modelo **{best_model.upper()}** fue seleccionado como el mejor y es el utilizado para inferencias en vivo.",
              f"🏆 The model **{best_model.upper()}** was selected as the best and is the one used for live inferences.",
              f"🏆 O modelo **{best_model.upper()}** foi selecionado como o melhor e é o usado para inferências ao vivo."))

    # Convertir JSON a DataFrame para mostrar
    records = []
    for m_name, details in model_details.items():
        records.append({
            t('Modelo', 'Model', 'Modelo'): f"⭐ {m_name.upper()}" if m_name == best_model else m_name.upper(),
            t('Accuracy', 'Accuracy', 'Acurácia'): f"{details.get('accuracy', 0)*100:.2f}%",
            t('F1 Score', 'F1 Score', 'Pontuação F1'): f"{details.get('f1', 0):.4f}",
            t('Tiempo de Entr. (s)', 'Training Time (s)', 'Tempo de Treino (s)'): f"{details.get('time_s', 0):.1f}"
        })
    df_eval = pd.DataFrame(records)
    
    st.table(df_eval)

    with st.expander(t("Ver Hiperparámetros Óptimos", "View Optimal Hyperparameters", "Ver Hiperparâmetros Ótimos")):
        for m_name, details in model_details.items():
            params = details.get('params', {})
            if params:
                st.write(f"**{m_name.upper()}**:")
                st.code(json.dumps(params, indent=2))
            else:
                st.write(t(f"**{m_name.upper()}**: Default", f"**{m_name.upper()}**: Default", f"**{m_name.upper()}**: Padrão"))

    if sport == 'futbol':
        try:
            with open('model_results_futbol_extras.json', 'r') as f:
                futbol_extras = json.load(f)
            
            for extra_key, extra_data in futbol_extras.items():
                st.write("")
                extra_title = t("⚽ Fútbol - Goles Over/Under 2.5", "⚽ Football - Goals Over/Under 2.5", "⚽ Futebol - Gols Over/Under 2.5") if extra_key == "futbol_over25" else t("⚽ Fútbol - Ambos Anotan (BTTS)", "⚽ Football - Both Teams to Score (BTTS)", "⚽ Futebol - Ambas Marcam (BTTS)")
                st.markdown(f"### {extra_title}")
                
                best_model_extra = extra_data['best_model']
                model_details_extra = extra_data['model_details']
                
                st.info(t(f"🏆 El modelo **{best_model_extra.upper()}** fue seleccionado como el mejor y es el utilizado para inferencias en vivo.",
                          f"🏆 The model **{best_model_extra.upper()}** was selected as the best and is the one used for live inferences.",
                          f"🏆 O modelo **{best_model_extra.upper()}** foi seleccionado como o melhor e é o usado para inferências ao vivo."))
                
                records_extra = []
                for m_name, details in model_details_extra.items():
                    records_extra.append({
                        t('Modelo', 'Model', 'Modelo'): f"⭐ {m_name.upper()}" if m_name == best_model_extra else m_name.upper(),
                        t('Accuracy', 'Accuracy', 'Acurácia'): f"{details.get('accuracy', 0)*100:.2f}%",
                        t('F1 Score', 'F1 Score', 'Pontuação F1'): f"{details.get('f1', 0):.4f}",
                        t('Tiempo de Entr. (s)', 'Training Time (s)', 'Tempo de Treino (s)'): f"{details.get('time_s', 0):.1f}"
                    })
                df_eval_extra = pd.DataFrame(records_extra)
                st.table(df_eval_extra)
                
                with st.expander(t(f"Ver Hiperparámetros Óptimos ({extra_title})", f"View Optimal Hyperparameters ({extra_title})", f"Ver Hiperparâmetros Ótimos ({extra_title})")):
                    for m_name, details in model_details_extra.items():
                        params = details.get('params', {})
                        if params:
                            st.write(f"**{m_name.upper()}**:")
                            st.code(json.dumps(params, indent=2))
                        else:
                            st.write(t(f"**{m_name.upper()}**: Default", f"**{m_name.upper()}**: Default", f"**{m_name.upper()}**: Padrão"))
        except FileNotFoundError:
            st.warning(t("⏳ Los modelos extras de fútbol (Over 2.5 y BTTS) se están entrenando en este momento. Por favor, regresa más tarde para ver los resultados.",
                         "⏳ Extra football models (Over 2.5 and BTTS) are currently training. Please return later to see the results.",
                         "⏳ Modelos extras de futebol (Over 2.5 e BTTS) estão treinando no momento. Por favor, volte mais tarde para ver os resultados."))

    st.divider()
    
    st.write(t("### 📄 Exportar Reporte Consolidado", "### 📄 Export Consolidated Report", "### 📄 Exportar Relatório Consolidado"))
    st.write(t("Descarga un informe detallado con las métricas y configuraciones de todos los modelos entrenados.",
               "Download a detailed report with the metrics and configurations of all trained models.",
               "Baixe um relatório detalhado com as métricas e configurações de todos os modelos treinados."))
    
    report_bytes = generate_consolidated_report('model_results.json')
    if report_bytes:
        st.download_button(
            label=t("⬇️ Descargar Reporte Completo (Word)", "⬇️ Download Full Report (Word)", "⬇️ Baixar Relatório Completo (Word)"),
            data=report_bytes,
            file_name=f"Reporte_Definitivo_Modelos_{datetime.now().strftime('%Y%m%d')}.docx",
            mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        )

def page_competitions():
    """Página de competencias"""
    st.title(t("🏆 Competencias y Torneos", "🏆 Competitions and Tournaments", "🏆 Competições e Torneios"))

    # Datos de ejemplo
    competitions_data = [
        {
            'id': 1,
            'name': t('Liga Predictor Enero', 'January Predictor League', 'Liga Predictor Janeiro'),
            'description': t('Predice correctamente y gana premios semanales', 'Predict correctly and win weekly prizes', 'Preveja corretamente e ganhe prêmios semanais'),
            'participants': 245,
            'prize': '$1,000',
            'entry_fee': t('Gratis', 'Free', 'Grátis'),
            'deadline': '2024-01-31',
            'status': 'active',
            'prize_pool': 1000
        },
        {
            'id': 2,
            'name': t('Torneo Premium Elite', 'Premium Elite Tournament', 'Torneio Premium Elite'),
            'description': t('Exclusivo para suscriptores premium - Grandes premios', 'Exclusive for premium subscribers - Big prizes', 'Exclusivo para assinantes premium - Grandes prêmios'),
            'participants': 89,
            'prize': '$2,500',
            'entry_fee': '$10',
            'deadline': '2024-01-25',
            'status': 'active',
            'prize_pool': 2500
        },
        {
            'id': 3,
            'name': t('Desafío de 100 Predicciones', '100 Predictions Challenge', 'Desafio de 100 Previsões'),
            'description': t('Realiza 100 predicciones precisas y gana jackpot', 'Make 100 accurate predictions and win jackpot', 'Faça 100 previsões precisas e ganhe jackpot'),
            'participants': 156,
            'prize': '$500',
            'entry_fee': t('Gratis', 'Free', 'Grátis'),
            'deadline': '2024-02-15',
            'status': 'active',
            'prize_pool': 500
        }
    ]

    col1, col2 = st.columns([2, 1])

    with col1:
        st.subheader(t("Competencias Disponibles", "Available Competitions", "Competições Disponíveis"))

        for comp in competitions_data:
            with st.expander(f"🏆 {comp['name']}"):
                st.write(t(f"**Descripción:** {comp['description']}", f"**Description:** {comp['description']}", f"**Descrição:** {comp['description']}"))

                col_a, col_b, col_c = st.columns(3)

                with col_a:
                    st.metric(t("Participantes", "Participants", "Participantes"), comp['participants'])

                with col_b:
                    st.metric(t("Premio Total", "Total Prize", "Prêmio Total"), comp['prize'])

                with col_c:
                    st.metric(t("Entrada", "Entry", "Entrada"), comp['entry_fee'])

                st.write(t(f"📅 **Cierra:** {comp['deadline']}", f"📅 **Closes:** {comp['deadline']}", f"📅 **Fecha:** {comp['deadline']}"))

                col_x, col_y = st.columns(2)

                with col_x:
                    if comp['entry_fee'] == t('Gratis', 'Free', 'Grátis'):
                        if st.button(t("✅ Unirse", "✅ Join", "✅ Juntar-se"), key=f"join_{comp['id']}"):
                            st.session_state.selected_competitions.append(comp['id'])
                            st.success(t(f"¡Te has unido a {comp['name']}!", f"You have joined {comp['name']}!", f"Você se juntou a {comp['name']}!"))
                            st.balloons()
                    else:
                        if st.session_state.user_tier in ['premium', 'pro', 'elite']:
                            if st.button(t("✅ Unirse", "✅ Join", "✅ Juntar-se"), key=f"join_{comp['id']}"):
                                st.session_state.selected_competitions.append(comp['id'])
                                st.success(t(f"¡Te has unido a {comp['name']}!", f"You have joined {comp['name']}!", f"Você se juntou a {comp['name']}!"))
                                st.balloons()
                        else:
                            st.warning(t("💎 Requiere suscripción Premium", "💎 Requires Premium subscription", "💎 Requer assinatura Premium"))

                with col_y:
                    st.write(t("Status: ✅ Activa", "Status: ✅ Active", "Status: ✅ Ativa"))

    with col2:
        st.subheader(t("📊 Tu Ranking", "📊 Your Rank", "📊 Sua Classificação"))

        ranking_positions = [
            {'pos': 1, 'user': 'PredictorPro', 'score': 1850, 'acc': '78%'},
            {'pos': 2, 'user': 'SportGenius', 'score': 1820, 'acc': '76%'},
            {'pos': 3, 'user': 'BetMaster', 'score': 1780, 'acc': '74%'},
            {'pos': 4, 'user': 'DataAnalyst', 'score': 1750, 'acc': '72%'},
            {'pos': 5, 'user': 'GoldenPredictor', 'score': 1720, 'acc': '71%'},
            {'pos': 15, 'user': t('👤 Tú', '👤 You', '👤 Você'), 'score': 1520, 'acc': '68%'},
        ]

        for rank in ranking_positions:
            if "Tú" in rank['user'] or "You" in rank['user'] or "Você" in rank['user'] or rank['user'] == t('👤 Tú', '👤 You', '👤 Você'):
                st.markdown(
                    f"<div style='background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); "
                    f"color: white; padding: 10px; border-radius: 5px; margin: 5px 0;'>"
                    f"<b>🎯 {rank['pos']}. {rank['user']}</b> - {rank['score']} pts ({rank['acc']})"
                    f"</div>",
                    unsafe_allow_html=True
                )
            else:
                st.write(f"{rank['pos']}. {rank['user']} - {rank['score']} pts ({rank['acc']})")

def page_statistics():
    """Página de estadísticas avanzadas con gráficos y reportes"""
    st.title(t("📊 ESTADÍSTICAS DESCRIPTIVAS Y ANÁLISIS", "📊 DESCRIPTIVE STATISTICS AND ANALYSIS", "📊 ESTATÍSTICAS DESCRITIVAS E ANÁLISE"))

    user_stats = get_user_stats(st.session_state.user_id)
    predictions = get_user_predictions(st.session_state.user_id)

    # ============ TAB 1: RESUMEN ============
    tab1, tab2, tab3, tab4, tab5 = st.tabs(
        [t("📈 Resumen", "📈 Summary", "📈 Resumo"),
         t("📋 Tablas", "📋 Tables", "📋 Tabelas"),
         t("📊 Gráficos Avanzados", "📊 Advanced Charts", "📊 Gráficos Avançados"),
         t("📥 Reportes", "📥 Reports", "📥 Relatórios"),
         t("💡 Insights", "💡 Insights", "💡 Insights")]
    )

    with tab1:
        st.subheader(t("Resumen General", "General Summary", "Resumo Geral"))

        col1, col2, col3, col4 = st.columns(4)

        with col1:
            st.metric(
                t("Total Predicciones", "Total Predictions", "Total de Previsões"),
                user_stats['total_predictions'],
                delta=f"+{user_stats['total_predictions'] % 10}"
            )

        with col2:
            st.metric(
                t("Tasa de Precisión", "Accuracy Rate", "Taxa de Precisão"),
                f"{user_stats['accuracy_rate']:.1f}%",
                delta=f"+{user_stats['accuracy_rate'] - 55:.1f}%" if user_stats['accuracy_rate'] > 55 else "-2%"
            )

        with col3:
            st.metric(
                t("Confianza Promedio", "Average Confidence", "Confiança Média"),
                f"{user_stats['avg_confidence']:.2f}",
                delta="+0.05"
            )

        with col4:
            st.metric(
                t("Ranking Global", "Global Rank", "Classificação Global"),
                str(user_stats['rank']),
                delta=f"+5 " + t("posiciones", "positions", "posições")
            )

    # ============ TAB 2: TABLAS ESTADÍSTICAS ============
    with tab2:
        st.subheader(t("Estadísticas Descriptivas", "Descriptive Statistics", "Estatísticas Descritivas"))

        # Tabla de resumen
        col1, col2 = st.columns(2)

        with col1:
            st.write(t("**Resumen de Desempeño**", "**Performance Summary**", "**Resumo de Desempenho**"))
            summary_table = pd.DataFrame({
                t('Métrica', 'Metric', 'Métrica'): [
                    t('Total', 'Total', 'Total'),
                    t('Correctas', 'Correct', 'Corretas'),
                    t('Incorrectas', 'Incorrect', 'Incorretas'),
                    t('Precisión', 'Accuracy', 'Precisão'),
                    t('Confianza Prom.', 'Avg Confidence', 'Confiança Média'),
                    t('Ranking', 'Rank', 'Classificação')
                ],
                t('Valor', 'Value', 'Valor'): [
                    str(user_stats['total_predictions']),
                    str(user_stats['correct_predictions']),
                    str(user_stats['total_predictions'] - user_stats['correct_predictions']),
                    f"{user_stats['accuracy_rate']:.2f}%",
                    f"{user_stats['avg_confidence']:.2f}",
                    str(user_stats['rank'])
                ]
            })
            st.table(summary_table)

        with col2:
            st.write(t("**Estadísticas Descriptivas de Predicciones**", "**Descriptive Statistics of Predictions**", "**Estatísticas Descritivas de Previsões**"))
            stats_table = create_descriptive_stats_table(predictions)
            if not stats_table.empty:
                st.table(stats_table)
            else:
                st.info(t("No hay datos numéricos disponibles", "No numerical data available", "Não hay dados numéricos disponíveis"))

    # ============ TAB 3: GRÁFICOS AVANZADOS ============
    with tab3:
        st.subheader(t("Visualizaciones Avanzadas", "Advanced Visualizations", "Visualizações Avançadas"))

        if predictions and len(predictions) > 0:
            df = pd.DataFrame(predictions)

            # Gráfico 1: Distribución de Confianza
            if 'confidence_level' in df.columns:
                col1, col2 = st.columns(2)

                with col1:
                    confidence_data = df['confidence_level'].dropna().tolist()
                    fig_dist = create_distribution_plot(
                        confidence_data,
                        t("Distribución de Confianza (Histograma + Curva Normal)",
                          "Confidence Distribution (Histogram + Normal Curve)",
                          "Distribuição de Confiança (Histograma + Curva Normal)")
                    )
                    st.plotly_chart(fig_dist, use_container_width=True)

                with col2:
                    # Gráfico de CDF (Función de Distribución Acumulada)
                    sorted_data = np.sort(confidence_data)
                    cumulative = np.arange(1, len(sorted_data) + 1) / len(sorted_data)

                    fig_cdf = go.Figure()
                    fig_cdf.add_trace(go.Scatter(
                        x=sorted_data, y=cumulative,
                        mode='lines',
                        name='CDF',
                        line=dict(color='#667EEA', width=3)
                    ))
                    fig_cdf.update_layout(
                        title=t("Función de Distribución Acumulada (CDF)", "Cumulative Distribution Function (CDF)", "Função de Distribuição Acumulada (CDF)"),
                        xaxis_title=t("Confianza", "Confidence", "Confiança"),
                        yaxis_title=t("Probabilidad", "Probability", "Probabilidade"),
                        height=400
                    )
                    st.plotly_chart(fig_cdf, use_container_width=True)

            # Gráfico 2: Serie Temporal
            if 'created_at' in df.columns:
                st.subheader(t("Evolución en el Tiempo", "Evolution over Time", "Evolução no Tempo"))

                df['created_at'] = pd.to_datetime(df['created_at'], errors='coerce')
                df_sorted = df.sort_values('created_at').dropna(subset=['created_at'])

                if len(df_sorted) > 0:
                    dates = df_sorted['created_at'].tolist()
                    values = df_sorted.get('confidence_level', pd.Series(range(len(df_sorted)))).tolist()

                    fig_time = create_time_series_plot(
                        dates, values,
                        t("Evolución de Confianza en el Tiempo", "Confidence Evolution over Time", "Evolução da Confiança no Tempo")
                    )
                    st.plotly_chart(fig_time, use_container_width=True)

            # Gráfico 3: Por Status
            if 'prediction_status' in df.columns:
                st.subheader(t("Análisis por Status", "Analysis by Status", "Análise por Status"))

                status_counts = df['prediction_status'].value_counts()
                fig_status = px.bar(
                    x=status_counts.index,
                    y=status_counts.values,
                    title=t("Distribución por Status", "Distribution by Status", "Distribuição por Status"),
                    labels={'x': t('Status', 'Status', 'Status'), 'y': t('Cantidad', 'Count', 'Quantidade')},
                    color_discrete_sequence=['#667EEA']
                )
                fig_status.update_layout(height=400)
                st.plotly_chart(fig_status, use_container_width=True)

        else:
            st.info(t("No hay suficientes datos para mostrar gráficos", "Not enough data to display charts", "Não há dados suficientes para exibir gráficos"))

    # ============ TAB 4: REPORTES ============
    with tab4:
        st.subheader(t("📥 Generar Reportes", "📥 Generate Reports", "📥 Gerar Relatórios"))

        col1, col2 = st.columns(2)

        with col1:
            st.write(t("**Reporte en PDF**", "**PDF Report**", "**Relatório em PDF**"))
            if st.button(t("📄 Descargar PDF", "📄 Download PDF", "📄 Baixar PDF"), key="btn_dld_pdf"):
                pdf_bytes = generate_predictions_report_pdf(predictions)
                st.download_button(
                    label=t("Descargar PDF", "Download PDF", "Baixar PDF"),
                    data=pdf_bytes,
                    file_name=f"predicciones_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
                    mime="application/pdf"
                )

        with col2:
            st.write(t("**Reporte en Excel**", "**Excel Report**", "**Relatório em Excel**"))
            if st.button(t("📊 Descargar Excel", "📊 Download Excel", "📊 Baixar Excel"), key="btn_dld_excel"):
                excel_bytes = generate_predictions_report_excel(predictions)
                st.download_button(
                    label=t("Descargar Excel", "Download Excel", "Baixar Excel"),
                    data=excel_bytes,
                    file_name=f"predicciones_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

    # ============ TAB 5: INSIGHTS ============
    with tab5:
        st.subheader(t("💡 Insights y Recomendaciones", "💡 Insights and Recommendations", "💡 Insights e Recomendações"))

        accuracy = user_stats.get('accuracy_rate', 0)
        confidence = user_stats.get('avg_confidence', 0)

        col1, col2 = st.columns(2)

        with col1:
            st.write(t("**Análisis de Rendimiento**", "**Performance Analysis**", "**Análise de Desempenho**"))

            if accuracy >= 75:
                st.success(t("✅ **EXCELENTE** - Tu precisión es superior al 75%", "✅ **EXCELLENT** - Your accuracy is above 75%", "✅ **EXCELENTE** - Sua precisão é superior a 75%"))
                insight1 = t("Mantén tu estrategia actual, has encontrado un buen patrón.", "Maintain your current strategy, you have found a good pattern.", "Mantenha sua estratégia atual, você encontrou um bom padrão.")
            elif accuracy >= 60:
                st.info(t("⚠️ **BUENO** - Tu precisión está sobre el promedio", "⚠️ **GOOD** - Your accuracy is above average", "⚠️ **BOM** - Sua precisão está acima da média"))
                insight1 = t("Continúa mejorando, hay espacio para optimizar.", "Keep improving, there is room to optimize.", "Continue melhorando, há espaço para otimizar.")
            else:
                st.warning(t("⚠️ **DESARROLLO** - Tu precisión está por debajo del 60%", "⚠️ **DEVELOPMENT** - Your accuracy is below 60%", "⚠️ **DESENVOLVIMENTO** - Sua precisão está abaixo de 60%"))
                insight1 = t("Revisa tu metodología de análisis.", "Review your analysis methodology.", "Revise sua metodologia de análise.")

            st.write(insight1)

        with col2:
            st.write(t("**Análisis de Confianza**", "**Confidence Analysis**", "**Análise de Confiança**"))

            if confidence >= 0.75:
                st.success(t("✅ Confianza bien calibrada (>0.75)", "✅ Well-calibrated confidence (>0.75)", "✅ Confiança bem calibrada (>0.75)"))
                insight2 = t("Tu confianza es realista con tu precisión.", "Your confidence is realistic with your accuracy.", "Sua confiança é realista com sua precisão.")
            elif confidence >= 0.6:
                st.info(t("⚠️ Confianza moderada (0.60-0.75)", "⚠️ Moderate confidence (0.60-0.75)", "⚠️ Confiança moderada (0.60-0.75)"))
                insight2 = t("Considera si tu confianza refleja tu precisión.", "Consider whether your confidence reflects your accuracy.", "Considere se sua confiança reflete sua precisão.")
            else:
                st.warning(t("⚠️ Confianza baja (<0.60)", "⚠️ Low confidence (<0.60)", "⚠️ Confiança baixa (<0.60)"))
                insight2 = t("Aumenta confianza cuando tengas análisis sólidos.", "Increase confidence when you have solid analysis.", "Aumente a confiança quando tiver análises sólidas.")

            st.write(insight2)

        st.divider()

        st.write(t("**Recomendaciones Personalizadas**", "**Personalized Recommendations**", "**Recomendações Personalizadas**"))
        recommendations = [
            t("🔍 Analiza tus predicciones incorrectas para identificar patrones", "🔍 Analyze your incorrect predictions to identify patterns", "🔍 Analise suas previsões incorretas para identificar padrões"),
            t("📚 Aumenta confianza solo cuando tengas datos sólidos de respaldo", "📚 Increase confidence only when you have solid backup data", "📚 Aumente a confiança apenas quando tiver dados de apoio sólidos"),
            t("🎯 Diversifica tus predicciones entre diferentes deportes y ligas", "🎯 Diversify your predictions across different sports and leagues", "🎯 Diversifique suas previsões entre diferentes esportes e ligas"),
            t("📊 Mantén un registro detallado de tus análisis y resultados", "📊 Maintain a detailed record of your analyses and results", "📊 Mantenha um registro detalhado de suas análises e resultados"),
            t("⏱️ Revisa tu rendimiento regularmente (semanal/mensual)", "⏱️ Review your performance regularly (weekly/monthly)", "⏱️ Revise seu desempenho regularmente (semanal/mensal)"),
            t("💡 Considera variables externas: lesiones, clima, forma del equipo", "💡 Consider external variables: injuries, weather, team form", "💡 Considere variáveis externas: lesões, clima, forma da equipe")
        ]

        for rec in recommendations:
            st.write(rec)

def page_alerts():
    """Página de alertas"""
    st.title(t("🔔 Alertas y Notificaciones", "🔔 Alerts and Notifications", "🔔 Alertas e Notificações"))

    alerts_data = [
        {
            'type': 'partido',
            'icon': '⚽',
            'title': t('Nuevo partido disponible', 'New match available', 'Novo jogo disponível'),
            'message': t('Real Madrid vs Barcelona - Predicción IA lista', 'Real Madrid vs Barcelona - AI Prediction ready', 'Real Madrid vs Barcelona - Previsão de IA pronta'),
            'time': t('Hace 2 horas', '2 hours ago', 'Há 2 horas'),
            'read': False
        },
        {
            'type': 'prediccion',
            'icon': '🤖',
            'title': t('Actualización de predicción IA', 'AI prediction update', 'Atualização de previsão de IA'),
            'message': t('Nueva predicción para Manchester City vs Liverpool', 'New prediction for Manchester City vs Liverpool', 'Nova previsão para Manchester City vs Liverpool'),
            'time': t('Hace 5 horas', '5 hours ago', 'Há 5 horas'),
            'read': True
        },
        {
            'type': 'competencia',
            'icon': '🏆',
            'title': t('Competencia finaliza pronto', 'Competition ending soon', 'Competição terminando em breve'),
            'message': t('Liga Predictor Enero cierra en 2 días', 'January Predictor League closes in 2 days', 'Liga Predictor Janeiro fecha em 2 dias'),
            'time': t('Hace 1 día', '1 day ago', 'Há 1 dia'),
            'read': False
        },
        {
            'type': 'ranking',
            'icon': '📈',
            'title': t('Cambio en tu ranking', 'Rank change', 'Mudança na sua classificação'),
            'message': t('Subiste 5 posiciones en el ranking global', 'You rose 5 positions in the global ranking', 'Você subiu 5 posições no ranking global'),
            'time': t('Hace 1 día', '1 day ago', 'Há 1 dia'),
            'read': True
        }
    ]

    col1, col2 = st.columns([3, 1])

    with col1:
        st.subheader(t("Notificaciones Recientes", "Recent Notifications", "Notificações Recentes"))

        for alert in alerts_data:
            status_icon = "📩" if alert['read'] else "📧"
            bg_color = ("#1e293b" if st.session_state.theme == "dark" else "#f1f5f9") if alert['read'] else ("#4c3a2b" if st.session_state.theme == "dark" else "#ffedd5")
            text_color = "#f8fafc" if st.session_state.theme == "dark" else "#1e293b"

            col_alert = st.columns([1, 10, 1])

            with col_alert[0]:
                st.write(alert['icon'])

            with col_alert[1]:
                st.markdown(
                    f"<div style='background: {bg_color}; padding: 15px; border-radius: 5px; color: {text_color};'>"
                    f"<b>{alert['title']}</b><br>"
                    f"{alert['message']}<br>"
                    f"<small>{alert['time']}</small>"
                    f"</div>",
                    unsafe_allow_html=True
                )

            with col_alert[2]:
                st.write(status_icon)

    with col2:
        st.subheader(t("Configurar Alertas", "Configure Alerts", "Configurar Alertas"))

        st.checkbox(t("⚽ Nuevos partidos", "⚽ New matches", "⚽ Novos jogos"), value=True, key="chk_alert_matches")
        st.checkbox(t("🤖 Predicciones IA", "🤖 AI Predictions", "🤖 Previsões de IA"), value=True, key="chk_alert_ai")
        st.checkbox(t("🏆 Competencias", "🏆 Competitions", "🏆 Competições"), value=True, key="chk_alert_comps")
        st.checkbox(t("📈 Cambios ranking", "📈 Rank changes", "📈 Mudanças de classificação"), value=True, key="chk_alert_rank")
        st.checkbox(t("💎 Ofertas Premium", "💎 Premium offers", "💎 Ofertas Premium"), value=False, key="chk_alert_premium")

        if st.button(t("Guardar configuración", "Save configuration", "Salvar configuração"), key="btn_save_alerts_config"):
            st.success(t("✅ Configuración guardada", "✅ Configuration saved", "✅ Configuração salva"))



def page_premium():
    """Página de suscripción premium"""
    st.title(t("💎 SportsPredict Pro Premium", "💎 SportsPredict Pro Premium", "💎 SportsPredict Pro Premium"))

    col1, col2, col3 = st.columns(3)

    with col1:
        st.subheader(t("📦 Plan Free", "📦 Free Plan", "📦 Plano Free"))

        st.markdown(t("""
        ✅ Predicciones IA básicas
        ✅ Competencias gratuitas
        ✅ Estadísticas básicas
        ✅ Hasta 50 predicciones/mes

        ---

        **Precio:** Gratis
        """, """
        ✅ Basic AI predictions
        ✅ Free competitions
        ✅ Basic statistics
        ✅ Up to 50 predictions/month

        ---

        **Price:** Free
        """, """
        ✅ Previsões básicas de IA
        ✅ Competições gratuitas
        ✅ Estatísticas básicas
        ✅ Até 50 previsões/mês

        ---

        **Preço:** Grátis
        """))

        if st.session_state.user_tier != "free":
            st.info(t("Tu plan actual: Premium", "Your current plan: Premium", "Seu plano atual: Premium"))
        else:
            st.success(t("Tu plan actual", "Your current plan", "Seu plano atual"))

    with col2:
        st.subheader(t("🌟 Plan Pro", "🌟 Pro Plan", "🌟 Plano Pro"))

        st.markdown(t("""
        ✅ Todas las del plan Free
        ✅ Predicciones IA avanzadas
        ✅ Acceso a competencias premium
        ✅ Análisis detallados
        ✅ Reportes PDF personalizados
        ✅ Alertas ilimitadas
        ✅ Hasta 500 predicciones/mes

        ---

        **Precio:** $4.99/mes
        *(o $49.99/año)*
        """, """
        ✅ All Free plan features
        ✅ Advanced AI predictions
        ✅ Access to premium competitions
        ✅ Detailed analysis
        ✅ Customized PDF reports
        ✅ Unlimited alerts
        ✅ Up to 500 predictions/month

        ---

        **Price:** $4.99/month
        *(or $49.99/year)*
        """, """
        ✅ Todos os recursos do plano Free
        ✅ Previsões avançadas de IA
        ✅ Acesso a competições premium
        ✅ Análises detalhadas
        ✅ Relatórios PDF personalizados
        ✅ Alertas ilimitados
        ✅ Até 500 previsões/mês

        ---

        **Preço:** $4.99/mês
        *(ou $49.99/ano)*
        """))

        if st.button(t("🚀 Suscribirse a Pro", "🚀 Subscribe to Pro", "🚀 Assinar Pro"), key="btn_pro"):
            st.session_state.user_tier = "pro"
            st.success(t("¡Bienvenido a SportsPredict Pro!", "Welcome to SportsPredict Pro!", "Bem-vindo ao SportsPredict Pro!"))
            st.balloons()

    with col3:
        st.subheader(t("👑 Plan Elite", "👑 Elite Plan", "👑 Plano Elite"))

        st.markdown(t("""
        ✅ Todas las del plan Pro
        ✅ Acceso a análisis IA en tiempo real
        ✅ Consultas personalizadas con expertos
        ✅ Datos históricos completos
        ✅ Predicciones ilimitadas
        ✅ Prioridad en competencias
        ✅ Descuentos en apuestas

        ---

        **Precio:** $9.99/mes
        *(o $99.99/año)*
        """, """
        ✅ All Pro plan features
        ✅ Access to real-time AI analysis
        ✅ Personalized expert consultations
        ✅ Complete historical data
        ✅ Unlimited predictions
        ✅ Priority in competitions
        ✅ Betting discounts

        ---

        **Price:** $9.99/month
        *(or $99.99/year)*
        """, """
        ✅ Todos os recursos do plano Pro
        ✅ Acesso a análises de IA em tempo real
        ✅ Consultas personalizadas com especialistas
        ✅ Dados históricos completos
        ✅ Previsões ilimitadas
        ✅ Prioridade em competições
        ✅ Descontos em apostas

        ---

        **Preço:** $9.99/mês
        *(ou $99.99/ano)*
        """))

        if st.button(t("👑 Suscribirse a Elite", "👑 Subscribe to Elite", "👑 Assinar Elite"), key="btn_elite"):
            st.session_state.user_tier = "elite"
            st.success(t("¡Bienvenido a SportsPredict Elite!", "Welcome to SportsPredict Elite!", "Bem-vindo ao SportsPredict Elite!"))
            st.balloons()

    st.divider()

    st.subheader(t("Beneficios Adicionales", "Additional Benefits", "Benefícios Adicionais"))

    benefits = pd.DataFrame({
        t('Característica', 'Feature', 'Recurso'): [
            t('Predicciones IA básicas', 'Basic AI Predictions', 'Previsões básicas de IA'),
            t('Predicciones IA avanzadas', 'Advanced AI Predictions', 'Previsões avançadas de IA'),
            t('Acceso a competencias', 'Competitions Access', 'Acesso a competições'),
            t('Reportes PDF', 'PDF Reports', 'Relatórios PDF'),
            t('Alertas personalizadas', 'Personalized Alerts', 'Alertas personalizadas'),
            t('Análisis histórico', 'Historical Analysis', 'Análise histórica'),
            t('Soporte prioritario', 'Priority Support', 'Suporte prioritário'),
            t('Consulta con expertos', 'Expert Consultation', 'Consulta com especialistas')
        ],
        'Free': ['✅', '❌', '✅', '❌', '❌', '❌', '❌', '❌'],
        'Pro': ['✅', '✅', '✅', '✅', '✅', '✅', '❌', '❌'],
        'Elite': ['✅', '✅', '✅', '✅', '✅', '✅', '✅', '✅']
    })

    st.table(benefits)



def check_login():
    if 'logged_in' not in st.session_state:
        st.session_state.logged_in = False
    
    if st.session_state.logged_in:
        return True

    set_sidebar_visibility(False)
        
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.title("🔐 SportsPredict Pro")
        
        tab1, tab2 = st.tabs([
            t("Iniciar Sesión", "Log In", "Iniciar Sessão"),
            t("Registrarse", "Register", "Cadastrar-se")
        ])
        
        with tab1:
            with st.form("login_form"):
                username = st.text_input(t("Usuario", "Username", "Usuário"))
                password = st.text_input(t("Contraseña", "Password", "Senha"), type="password")
                submit_login = st.form_submit_button(t("Ingresar", "Log In", "Entrar"))
                
                if submit_login:
                    user = authenticate_user(username, password)
                    if user:
                        st.session_state.logged_in = True
                        st.session_state.user_id = user.id
                        st.session_state.user_tier = user.plan
                        st.session_state.user_role = getattr(user, "role", "USER")
                        st.success(t(f"¡Bienvenido de nuevo, {username}!", f"Welcome back, {username}!", f"Bem-vindo de volta, {username}!"))
                        st.rerun()
                    else:
                        st.error(t("Credenciales inválidas", "Invalid credentials", "Credenciais inválidas"))

        with tab2:
            with st.form("register_form"):
                new_username = st.text_input(t("Nuevo Usuario", "New Username", "Novo Usuário"))
                new_password = st.text_input(t("Contraseña", "Password", "Senha"), type="password")
                plan_choice = st.selectbox(t("Plan Inicial", "Initial Plan", "Plano Inicial"), ["free", "pro", "elite"])
                submit_register = st.form_submit_button(t("Registrarse", "Register", "Cadastrar-se"))
                
                if submit_register:
                    try:
                        new_user = create_user(new_username, new_password, plan_choice, "USER")
                        st.success(t("Usuario creado exitosamente. Por favor, inicia sesión.", "User created successfully. Please log in.", "Usuário criado com sucesso. Por favor, faça o login."))
                    except Exception as e:
                        st.error(t(f"Error creando usuario (puede que ya exista): {e}", f"Error creating user (it might already exist): {e}", f"Erro ao criar usuário (pode ser que já exista): {e}"))

    return False

def logout():
    """Cierra la sesión actual y limpia el estado relacionado al usuario."""
    keys_to_clear = [
        'logged_in',
        'user_id',
        'user_tier',
        'user_role',
        'favorite_teams',
        'selected_competitions',
        'api_matches',
        'last_api_date'
    ]
    for key in keys_to_clear:
        if key in st.session_state:
            del st.session_state[key]
    st.rerun()

def set_sidebar_visibility(visible: bool):
    """Muestra u oculta el sidebar nativo de Streamlit y configura sus colores según el tema."""
    display = "block" if visible else "none"
    visibility = "visible" if visible else "hidden"
    width = "21rem" if visible else "0rem"
    margin_left = "0" if not visible else ""

    theme = st.session_state.get('theme', 'light')
    sidebar_bg = "#0b0f19" if theme == "dark" else "#f8fafc"
    sidebar_fg = "#f8fafc" if theme == "dark" else "#1e293b"

    components.html(
        f"""
        <script>
        const doc = window.parent.document;
        const sidebar = doc.querySelector('section[data-testid="stSidebar"]');
        const openBtn = doc.querySelector('button[data-testid="collapsedControl"]');
        const allButtons = Array.from(doc.querySelectorAll('button'));
        const toggleButtons = allButtons.filter(btn => {{
            const label = btn.getAttribute('aria-label') || '';
            return label.includes('sidebar') || label.includes('Sidebar');
        }});
        const main = doc.querySelector('section[data-testid="stMain"]');

        if (sidebar) {{
            sidebar.style.display = '{display}';
            sidebar.style.visibility = '{visibility}';
            sidebar.style.minWidth = '{width}';
            sidebar.style.maxWidth = '{width}';
            sidebar.style.width = '{width}';
            
            // Forzar color de fondo del sidebar principal
            sidebar.style.setProperty('background-color', '{sidebar_bg}', 'important');
            
            // Hacer transparentes todos los contenedores descendientes (excepto botones e inputs)
            // para que no tapen el color de fondo principal.
            const descendants = sidebar.querySelectorAll('*');
            descendants.forEach(el => {{
                const tag = el.tagName.toLowerCase();
                if (tag !== 'button' && tag !== 'input' && tag !== 'select' && tag !== 'textarea' && !el.classList.contains('stButton')) {{
                    el.style.setProperty('background-color', 'transparent', 'important');
                }}
            }});
            
            // Forzar color de texto de todos los elementos dentro del sidebar
            const textElements = sidebar.querySelectorAll('span, p, h1, h2, h3, h4, h5, h6, li, label, a');
            textElements.forEach(el => {{
                el.style.setProperty('color', '{sidebar_fg}', 'important');
            }});
        }}

        if (openBtn) {{
            openBtn.style.display = '{display}';
            openBtn.style.visibility = '{visibility}';
        }}

        toggleButtons.forEach(btn => {{
            btn.style.display = '{display}';
            btn.style.visibility = '{visibility}';
        }});

        if (main && '{margin_left}') {{
            main.style.marginLeft = '{margin_left}';
        }}
        </script>
        """,
        height=0,
        width=0,
    )

def page_create_prediction():
    """Página para crear predicciones manuales"""
    st.title(t("➕ Crear Predicción Manual", "➕ Create Manual Prediction", "➕ Criar Previsão Manual"))
    st.markdown(t(
        "Ingresa un partido para que la Inteligencia Artificial analice quién ganará basándose en su historial estadístico.",
        "Enter a match for the Artificial Intelligence to analyze who will win based on its statistical history.",
        "Insira um jogo para que a Inteligência Artificial analise quem vencerá com base no seu histórico estatístico."
    ))
    
    col1, col2 = st.columns([1, 2])
    with col1:
        sport = st.selectbox(t("Selecciona el Deporte", "Select Sport", "Selecione o Esporte"), ["Futbol", "NBA", "MLB"])
        
    st.divider()
    
    teams = get_teams_by_sport(sport)
    
    col1, col2, col3 = st.columns([2, 1, 2])
    
    with col1:
        home_team = st.selectbox(t("Equipo Local", "Home Team", "Equipe Local") if sport != 'MLB' else t("Equipo 1", "Team 1", "Equipe 1"), teams, key="home_select")
        pitcher1 = None
        if sport == 'MLB':
            pitchers = get_pitchers()
            pitcher1 = st.selectbox(t("Lanzador Abridor 1", "Starting Pitcher 1", "Arremessador Inicial 1"), pitchers, key="p1_select")
            
    with col3:
        # Default al segundo equipo para evitar que local y visitante sean el mismo al inicio
        default_away = teams[1] if len(teams) > 1 else teams[0]
        away_team = st.selectbox(t("Equipo Visitante", "Away Team", "Equipe Visitante") if sport != 'MLB' else t("Equipo 2", "Team 2", "Equipe 2"), teams, index=teams.index(default_away) if default_away in teams else 0, key="away_select")
        pitcher2 = None
        if sport == 'MLB':
            pitcher2 = st.selectbox(t("Lanzador Abridor 2", "Starting Pitcher 2", "Arremessador Inicial 2"), pitchers, key="p2_select")
            
    with col2:
        st.markdown("<h2 style='text-align: center; margin-top: 30px;'>VS</h2>", unsafe_allow_html=True)
        
    date_match = st.date_input(t("Fecha del Partido", "Match Date", "Data do Jogo"))
    
    if st.button(t("🔮 Generar Predicción IA", "🔮 Generate AI Prediction", "🔮 Gerar Previsão de IA"), type="primary", use_container_width=True):
        if home_team == away_team:
            st.error(t("El equipo local y visitante no pueden ser el mismo.", "Home and away teams cannot be the same.", "O time da casa e o visitante não podem ser o mesmo."))
            return
            
        with st.spinner(t("Analizando historial y ejecutando modelos...", "Analyzing history and running models...", "Analisando histórico e executando modelos...")):
            try:
                if sport == 'Futbol':
                    probs = predict_football(home_team, away_team)
                    # Probabilidades principales: local, empate, visitante
                    main_probs = {k: probs[k] for k in ['visitante', 'empate', 'local']}
                    labels = [t('Visitante', 'Away', 'Visitante'), t('Empate', 'Draw', 'Empate'), t('Local', 'Home', 'Local')]
                    values = [main_probs['visitante']*100, main_probs['empate']*100, main_probs['local']*100]
                    pred_label = max(main_probs, key=main_probs.get)
                    
                    over_str = t("MÁS de 2.5", "OVER 2.5", "MAIS de 2.5") if probs.get('over25', 0) > 0.5 else t("MENOS de 2.5", "UNDER 2.5", "MENOS de 2.5")
                    btts_str = t("SÍ", "YES", "SIM") if probs.get('btts', 0) > 0.5 else t("NO", "NO", "NÃO")
                    pred_label_db = f"{pred_label.upper()} | Over2.5: {over_str} ({probs.get('over25',0)*100:.1f}%) | BTTS: {btts_str} ({probs.get('btts',0)*100:.1f}%)"
                elif sport == 'NBA':
                    probs = predict_nba(home_team, away_team)
                    labels = [t('Visitante', 'Away', 'Visitante'), t('Local', 'Home', 'Local')]
                    values = [probs['visitante']*100, probs['local']*100]
                    pred_label = max(probs, key=probs.get)
                elif sport == 'MLB':
                    probs = predict_mlb(home_team, away_team, pitcher1, pitcher2)
                    labels = [t('Equipo 2', 'Team 2', 'Equipe 2'), t('Equipo 1', 'Team 1', 'Equipe 1')]
                    values = [probs['visitante']*100, probs['local']*100]
                    pred_label = max(probs, key=probs.get)
                    pred_label_db = pred_label.upper()
                    
                st.success(t(f"¡Predicción generada! El modelo se inclina por: **{pred_label.upper()}**",
                             f"Prediction generated! The model leans towards: **{pred_label.upper()}**",
                             f"Previsão gerada! O modelo se inclina por: **{pred_label.upper()}**"))
                if sport == 'Futbol':
                    st.info(t(f"⚽ **Goles:** {over_str} ({probs.get('over25',0)*100:.1f}%) | **Ambos Anotan:** {btts_str} ({probs.get('btts',0)*100:.1f}%)",
                              f"⚽ **Goals:** {over_str} ({probs.get('over25',0)*100:.1f}%) | **Both Teams to Score:** {btts_str} ({probs.get('btts',0)*100:.1f}%)",
                              f"⚽ **Gols:** {over_str} ({probs.get('over25',0)*100:.1f}%) | **Ambas Marcam:** {btts_str} ({probs.get('btts',0)*100:.1f}%)"))
                
                # Visualización
                fig = go.Figure(data=[go.Pie(labels=labels, values=values, hole=.5, marker_colors=['#f5576c', '#aaaaaa', '#667eea'])])
                fig.update_layout(title_text=t("Probabilidades de Victoria", "Win Probabilities", "Probabilidades de Vitória"))
                st.plotly_chart(fig, use_container_width=True)
                
                # Guardar el partido futuro
                save_future_match(sport, date_match.strftime("%Y-%m-%d"), home_team, away_team, pitcher1, pitcher2, probs, pred_label_db)
                
                # Guardar automáticamente la predicción en el perfil si es usuario Elite
                if st.session_state.user_tier == 'elite':
                    try:
                        prisma_save_prediction(
                            user_id=st.session_state.user_id,
                            match_id=int(time.time()), # Generar un ID dummy único para el partido
                            predicted_winner=pred_label_db,
                            confidence=max(values)/100.0,
                            is_manual=True,
                            ai_data=json.dumps({"prediccion_ia": pred_label, "probs": probs})
                        )
                        st.success(t("🌟 (Elite) Predicción guardada en tu perfil personal automáticamente.",
                                     "🌟 (Elite) Prediction automatically saved to your personal profile.",
                                     "🌟 (Elite) Previsão salva automaticamente no seu perfil pessoal."))
                    except Exception as e:
                        st.warning(t(f"No se pudo guardar en el perfil Elite: {e}", f"Could not save to Elite profile: {e}", f"Não foi possível salvar no perfil Elite: {e}"))

                st.info(t("Este partido ha sido guardado y aparecerá en el Dashboard.", "This match has been saved and will appear on the Dashboard.", "Este jogo foi salvo e aparecerá no Dashboard."))
            except Exception as e:
                st.error(t(f"Error generando predicción: {str(e)}", f"Error generating prediction: {str(e)}", f"Erro ao gerar previsão: {str(e)}"))



def main():
    """Función principal"""
    
    if not check_login():
        return

    # Inyectar estilos en cada corrida de página para evitar que st.navigation limpie el DOM del CSS
    inject_theme_css()
    set_sidebar_visibility(True)

    # Sidebar
    # Menú de navegación moderno
    pages = [
        st.Page(page_dashboard, title=t("Dashboard", "Dashboard", "Dashboard"), icon="🏠"),
        st.Page(page_create_prediction, title=t("Crear Predicción", "Create Prediction", "Criar Previsão"), icon="➕"),
        st.Page(page_my_predictions, title=t("Mis Predicciones", "My Predictions", "Minhas Previsões"), icon="📊"),
        st.Page(page_ai_predictions, title=t("Predicciones IA", "AI Predictions", "Previsões IA"), icon="🤖"),
        st.Page(page_competitions, title=t("Competencias", "Competitions", "Competições"), icon="🏆"),
        st.Page(page_statistics, title=t("Estadísticas", "Statistics", "Estatísticas"), icon="📈"),
        st.Page(page_alerts, title=t("Alertas", "Alerts", "Alertas"), icon="🔔"),
        st.Page(page_premium, title=t("Premium", "Premium", "Premium"), icon="💎")
    ]

    pg = st.navigation(pages, position="hidden")

    with st.sidebar:
        # 1. Logo de la página arriba de todo
        st.markdown(
            """
            <div class="sidebar-logo-container">
                <span class="sidebar-logo-icon">⚽</span>
                <span class="sidebar-logo-text">SportsPredict <span class="pro-tag">Pro</span></span>
            </div>
            """, 
            unsafe_allow_html=True
        )
        
        # 2. Etiqueta estática (sin desplegable)
        st.markdown(
            f"<div class='sidebar-nav-label'>{t('Navegación Principal', 'Main Navigation', 'Navegação Principal')}</div>", 
            unsafe_allow_html=True
        )
        
        # 3. Tarjetas de navegación (st.page_link)
        for page in pages:
            st.page_link(page, label=page.title, icon=page.icon, use_container_width=True)
            
        st.divider()
        
        # 4. Información del usuario (Tarjeta moderna)
        user_id = st.session_state.user_id or ""
        st.markdown(
            f"""
            <div class="user-info-card">
                <div class="user-info-item">👤 <b>{t('Usuario', 'User', 'Usuário')}:</b> <code>{user_id[:12]}...</code></div>
                <div class="user-info-item">💳 <b>{t('Plan', 'Plan', 'Plano')}:</b> <span class="premium-badge" style="padding: 2px 8px; font-size: 0.75rem;">{st.session_state.user_tier.upper()}</span></div>
                <div class="user-info-item">🛡️ <b>{t('Rol', 'Role', 'Função')}:</b> <code>{st.session_state.user_role.upper()}</code></div>
            </div>
            """,
            unsafe_allow_html=True
        )
        
        # 5. Botón de Cerrar Sesión
        if st.button(t("🚪 Cerrar Sesión", "🚪 Log Out", "🚪 Sair"), use_container_width=True, type="secondary"):
            logout()
            
        st.divider()
        
        # 6. Información adicional / Pie de página
        st.write(t("### 📱 Información", "### 📱 Information", "### 📱 Informações"))
        st.info(t(
            "**SportsPredict Pro v1.0**\n\nPlataforma de predicciones deportivas con IA\n\n🔗 [Sitio Web](https://example.com)\n📧 [Contacto](mailto:info@example.com)",
            "**SportsPredict Pro v1.0**\n\nAI sports prediction platform\n\n🔗 [Website](https://example.com)\n📧 [Contact](mailto:info@example.com)",
            "**SportsPredict Pro v1.0**\n\nPlataforma de previsões esportivas com IA\n\n🔗 [Sítio Web](https://example.com)\n📧 [Contato](mailto:info@example.com)"
        ))

    # Renderizar página seleccionada nativamente
    pg.run()

if __name__ == "__main__":
    main()
